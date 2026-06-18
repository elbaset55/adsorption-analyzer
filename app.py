import streamlit as st
import pandas as pd
import numpy as np
from scipy.optimize import curve_fit
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io, math

# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="AdsorpLab Pro", page_icon="🔬",
                   layout="wide", initial_sidebar_state="expanded")

# ═══════════════════════════════════════════════════════════════════════════════
#  SESSION DEFAULTS
# ═══════════════════════════════════════════════════════════════════════════════
for k, v in [("lang","en"),("dark",False)]:
    if k not in st.session_state: st.session_state[k] = v

DEMO_ISO = pd.DataFrame({'C0':[10,20,40,60,80,100,150,200],
                         'Absorbance':[0.12,0.22,0.38,0.50,0.60,0.68,0.78,0.84]})
DEMO_KIN = pd.DataFrame({'Time':[2,5,10,15,20,30,45,60,90,120],
                         'Absorbance':[0.70,0.58,0.44,0.35,0.28,0.22,0.17,0.14,0.12,0.11]})
for k, v in [("iso_data",DEMO_ISO.copy()),("kin_data",DEMO_KIN.copy())]:
    if k not in st.session_state: st.session_state[k] = v

# ═══════════════════════════════════════════════════════════════════════════════
#  TRANSLATIONS
# ═══════════════════════════════════════════════════════════════════════════════
TRANS = {
"en":{
  "app_title":"AdsorpLab Pro","app_subtitle":"Advanced Adsorption & Kinetics Analysis Platform",
  "app_badge":"✦ World-Class Research Tool","lang_btn":"🌐 العربية",
  "dark_btn":"☀️ Light","light_btn":"🌙 Dark",
  "cal_settings":"Calibration Settings","cal_method_label":"Concentration Method",
  "cal_beer":"Beer-Lambert","cal_curve":"Standard Curve",
  "epsilon":"ε (L/mol·cm)","path_len":"Path length b (cm)","molar_mass":"Molar mass (g/mol)",
  "slope":"Slope (m)","intercept":"Intercept",
  "visualization":"Visualization","show_residuals":"Residual Plots","marker_size":"Marker Size",
  "quick_actions":"Quick Actions","demo_iso":"🧪 Demo: Isotherms","demo_kin":"⏱ Demo: Kinetics",
  "tab_home":"🏠 Home","tab_cal":"⚗️ Calibration","tab_iso":"📊 Isotherms",
  "tab_kin":"⏱️ Kinetics","tab_thermo":"🌡️ Thermodynamics","tab_report":"📄 Report",
  "home_welcome":"Welcome to AdsorpLab Pro",
  "home_sub":"The most advanced adsorption analysis tool — built for researchers, designed for excellence.",
  "home_feat_title":"Platform Capabilities",
  "feat1_t":"4 Isotherm Models","feat1_d":"Langmuir, Freundlich, Temkin, D-R, Sips, BET",
  "feat2_t":"4 Kinetic Models","feat2_d":"PFO, PSO, Elovich, Weber-Morris + Arrhenius Ea",
  "feat3_t":"Thermodynamics","feat3_d":"ΔH°, ΔS°, ΔG° via Van't Hoff equation",
  "feat4_t":"Statistical Analysis","feat4_d":"R², RMSE, χ², confidence metrics",
  "feat5_t":"Linear Transforms","feat5_d":"Linearized isotherm plots for validation",
  "feat6_t":"Excel Reports","feat6_d":"Full multi-sheet scientific report export",
  "guide_title":"Quick Start Guide",
  "guide1":"⚗️  Set your calibration method in the sidebar (Beer-Lambert or Standard Curve).",
  "guide2":"📊  Go to Isotherms — load demo data or enter your own, select models and fit.",
  "guide3":"⏱️  Go to Kinetics — enter time-absorbance data and identify the best kinetic model.",
  "guide4":"🌡️  Enter Kc values at different temperatures for full thermodynamic analysis.",
  "guide5":"📄  Visit Report to download a complete Excel report of all results.",
  "model_ref":"Model Quick Reference",
  "session_title":"Session Status",
  "iso_status":"Isotherms","kin_status":"Kinetics","thermo_status":"Thermodynamics",
  "not_run":"Not analyzed yet","done":"✅ Analysis complete",
  "cal_header":"Calibration & Absorbance","cal_desc":"Configure your spectrophotometric calibration method",
  "live_preview":"🔍 Live Concentration Preview","test_abs":"Test absorbance value",
  "calc_conc":"Calculated Concentration","cal_curve_vis":"📈 Calibration Curve",
  "conc_mgl":"Concentration (mg/L)","absorbance":"Absorbance",
  "iso_header":"Equilibrium Isotherm Analysis",
  "iso_desc":"Langmuir · Freundlich · Temkin · D-R · Sips · BET",
  "vol":"Solution Volume V (L)","mass":"Adsorbent Mass m (g)",
  "models_fit":"Models to fit","data_input":"Data Input Method",
  "upload":"📁 Upload Excel/CSV","manual":"✏️ Manual Entry",
  "upload_hint_iso":"File must have columns: **C0** and **Absorbance**",
  "upload_hint_kin":"File must have columns: **Time** and **Absorbance**",
  "avg_removal":"Avg Removal","max_qe":"Max qₑ (mg/g)","max_ce":"Max Cₑ (mg/L)",
  "data_pts":"Data Points","model_ranking":"🏆 Model Ranking","best_fit":"BEST FIT",
  "parameters":"🔢 Parameters","computed_data":"📋 Computed Data",
  "iso_curves":"📈 Isotherm Fitting","residual_plot":"📉 Residuals",
  "linear_plots":"📐 Linear Transform Plots","r2_chart":"📊 Model R² Comparison",
  "auto_interp":"🤖 Auto Interpretation",
  "iso_min_pts":"📝 Enter at least 3 data points to begin.",
  "langmuir_linear":"Langmuir Linear (Cₑ/qₑ vs Cₑ)",
  "freundlich_linear":"Freundlich Linear (ln qₑ vs ln Cₑ)",
  "kin_header":"Kinetic Model Analysis",
  "kin_desc":"PFO · PSO · Elovich · Weber-Morris · Arrhenius Activation Energy",
  "c0_kin":"Initial Concentration C₀ (mg/L)",
  "qt_max":"qₜ at t_max","max_time":"Max Time","max_qt":"Max qₜ",
  "kin_curves":"📈 Kinetic Fitting","wm_plot":"🌀 Weber-Morris Plot",
  "arrhenius_title":"🔥 Activation Energy (Arrhenius)","arrhenius_desc":"Enter rate constants at different temperatures",
  "arrhenius_exp":"A = Ae × exp(−Ea/RT) — slope = −Ea/R",
  "num_arr_temps":"Number of temperatures","arr_temp":"T (°C)","arr_k":"k value",
  "ea_result":"Activation Energy Eₐ","arr_r2":"Arrhenius R²",
  "ea_interp_phys":"Physical adsorption (Eₐ < 40 kJ/mol)",
  "ea_interp_chem":"Chemical adsorption (Eₐ > 40 kJ/mol)",
  "kin_min_pts":"📝 Enter at least 4 data points to begin.",
  "thermo_header":"Thermodynamic Analysis","thermo_desc":"Van 't Hoff · ΔG° · ΔH° · ΔS°",
  "num_temps":"Number of temperature points","temp_pt":"T (°C)","kc_pt":"Kc / Kd",
  "vth_plot":"📈 Van 't Hoff Plot","dg_plot":"📉 ΔG° vs Temperature",
  "thermo_interp":"🔬 Thermodynamic Interpretation",
  "enthalpy":"ΔH° (Enthalpy)","entropy":"ΔS° (Entropy)","dg_at_t1":"ΔG° at T₁","vth_r2":"R²",
  "exo_title":"🔥 Exothermic","endo_title":"❄️ Endothermic",
  "exo_body":"Heat is released. Adsorption decreases with temperature.",
  "endo_body":"Heat is absorbed. Adsorption increases with temperature.",
  "entropy_pos":"📈 Increased Randomness","entropy_neg":"📉 Decreased Randomness",
  "entropy_pos_body":"ΔS° > 0 — increased disorder at solid-solution interface.",
  "entropy_neg_body":"ΔS° < 0 — more ordered interface.",
  "dg_table":"📋 ΔG° Summary","spontaneous":"✅ Spontaneous at all temperatures",
  "non_spontaneous":"⚠️ Non-spontaneous at some temperatures",
  "exo_desc":"Exo (<0) / Endo (>0)","entropy_desc":"Order change","linear_fit":"Fit quality",
  "duration":"Experiment duration","peak_kin":"Peak kinetic capacity",
  "valid_meas":"Valid measurements","mean_across":"Mean all points",
  "peak_ads":"Peak adsorption","max_eq_conc":"Max equilibrium conc.",
  "report_header":"Report & Export","report_desc":"Generate comprehensive Excel report",
  "exp_name":"Experiment / Sample ID","researcher":"Researcher Name","notes_label":"Notes",
  "export_includes":"📦 Export Includes","generate_btn":"📊 Generate Excel Report",
  "download_btn":"⬇️ Download .xlsx","report_success":"✅ Ready! Click to download.",
  "session_results":"📊 Session Results","no_results":"No results yet — run analysis first.",
  "iso_models_lbl":"📊 Isotherm Models","kin_models_lbl":"⏱️ Kinetic Models","thermo_lbl":"🌡️ Thermodynamics",
  "demo_loaded":"Demo data loaded!","error_cols_iso":"Need columns: C0 and Absorbance",
  "error_cols_kin":"Need columns: Time and Absorbance",
  "conc_method":"Concentration Method: ",
  "interp_best":"Best fitting model is","interp_r2":"with R² =",
  "interp_lang_fav":"Favorable adsorption (RL between 0-1)",
  "interp_lang_unfav":"Unfavorable adsorption (RL > 1)",
  "interp_lang_irrev":"Irreversible adsorption (RL = 0)",
  "interp_frnd_good":"Favorable heterogeneous adsorption (1/n < 1)",
  "interp_frnd_poor":"Poor adsorption conditions (1/n > 1)",
  "data_quality":"📊 Data Quality",
  "dq_cv":"Coeff. of Variation","dq_range":"qₑ Range","dq_pts":"Valid Points",
  "dq_mono":"Monotonic","dq_yes":"Yes ✅","dq_no":"No ⚠️",
},
"ar":{
  "app_title":"أدسوربلاب برو","app_subtitle":"منصة تحليل الأدمصاص والحركية المتقدمة",
  "app_badge":"✦ أداة بحثية عالمية المستوى","lang_btn":"🌐 English",
  "dark_btn":"☀️ فاتح","light_btn":"🌙 داكن",
  "cal_settings":"إعدادات المعايرة","cal_method_label":"طريقة حساب التركيز",
  "cal_beer":"بير-لامبرت","cal_curve":"منحنى معايرة قياسي",
  "epsilon":"ε (L/mol·cm)","path_len":"طول مسار الضوء b (cm)","molar_mass":"الكتلة المولية (g/mol)",
  "slope":"الميل (m)","intercept":"الجزء المقطوع",
  "visualization":"خيارات العرض","show_residuals":"رسم البواقي","marker_size":"حجم النقاط",
  "quick_actions":"إجراءات سريعة","demo_iso":"🧪 بيانات تجريبية: أيزوثيرم","demo_kin":"⏱ بيانات تجريبية: حركية",
  "tab_home":"🏠 الرئيسية","tab_cal":"⚗️ المعايرة","tab_iso":"📊 الأيزوثيرم",
  "tab_kin":"⏱️ الحركية","tab_thermo":"🌡️ الثرموديناميكا","tab_report":"📄 التقرير",
  "home_welcome":"مرحباً بك في أدسوربلاب برو",
  "home_sub":"أكثر أداة تحليل أدمصاص احترافية — مصممة للباحثين، بمستوى عالمي.",
  "home_feat_title":"إمكانيات المنصة",
  "feat1_t":"6 نماذج أيزوثيرم","feat1_d":"لانجمير، فرندليش، تمكين، D-R، Sips، BET",
  "feat2_t":"4 نماذج حركية","feat2_d":"PFO، PSO، إيلوفيتش، ويبر-موريس + طاقة التنشيط",
  "feat3_t":"الثرموديناميكا","feat3_d":"ΔH°، ΔS°، ΔG° عبر معادلة فانت هوف",
  "feat4_t":"التحليل الإحصائي","feat4_d":"R²، RMSE، χ²، مقاييس الثقة",
  "feat5_t":"الصور الخطية","feat5_d":"مخططات الأيزوثيرم المحوّلة للتحقق",
  "feat6_t":"تقارير Excel","feat6_d":"تقرير علمي متعدد الأوراق قابل للتصدير",
  "guide_title":"دليل البدء السريع",
  "guide1":"⚗️  اختر طريقة المعايرة في الشريط الجانبي (بير-لامبرت أو منحنى قياسي).",
  "guide2":"📊  اذهب إلى الأيزوثيرم — حمّل بيانات تجريبية أو أدخل بياناتك، واختر النماذج.",
  "guide3":"⏱️  اذهب إلى الحركية — أدخل بيانات الزمن والامتصاصية وحدد أفضل نموذج.",
  "guide4":"🌡️  أدخل قيم Kc عند درجات حرارة مختلفة للتحليل الثرموديناميكي الكامل.",
  "guide5":"📄  زر التقرير لتنزيل ملف Excel شامل بجميع النتائج.",
  "model_ref":"مرجع سريع للنماذج",
  "session_title":"حالة الجلسة",
  "iso_status":"الأيزوثيرم","kin_status":"الحركية","thermo_status":"الثرموديناميكا",
  "not_run":"لم يُحلَّل بعد","done":"✅ اكتمل التحليل",
  "cal_header":"الامتصاصية ومنحنى المعايرة","cal_desc":"اختر طريقة المعايرة الطيفية",
  "live_preview":"🔍 معاينة التركيز الآني","test_abs":"أدخل قيمة الامتصاصية",
  "calc_conc":"التركيز المحسوب","cal_curve_vis":"📈 منحنى المعايرة",
  "conc_mgl":"التركيز (mg/L)","absorbance":"الامتصاصية",
  "iso_header":"تحليل نماذج الأيزوثيرم",
  "iso_desc":"لانجمير · فرندليش · تمكين · D-R · Sips · BET",
  "vol":"حجم المحلول V (L)","mass":"كتلة المادة الماصة m (g)",
  "models_fit":"النماذج","data_input":"طريقة الإدخال",
  "upload":"📁 رفع ملف Excel/CSV","manual":"✏️ إدخال يدوي",
  "upload_hint_iso":"يجب أن يحتوي الملف على عمودي: **C0** و **Absorbance**",
  "upload_hint_kin":"يجب أن يحتوي الملف على عمودي: **Time** و **Absorbance**",
  "avg_removal":"متوسط الإزالة","max_qe":"أقصى qₑ","max_ce":"أقصى Cₑ",
  "data_pts":"عدد النقاط","model_ranking":"🏆 ترتيب النماذج","best_fit":"الأفضل",
  "parameters":"🔢 المعاملات","computed_data":"📋 البيانات المحسوبة",
  "iso_curves":"📈 منحنيات الأيزوثيرم","residual_plot":"📉 البواقي",
  "linear_plots":"📐 مخططات التحويل الخطي","r2_chart":"📊 مقارنة R²",
  "auto_interp":"🤖 التفسير التلقائي",
  "iso_min_pts":"📝 أدخل 3 نقاط على الأقل للبدء.",
  "langmuir_linear":"لانجمير الخطي (Cₑ/qₑ مقابل Cₑ)",
  "freundlich_linear":"فرندليش الخطي (ln qₑ مقابل ln Cₑ)",
  "kin_header":"تحليل النماذج الحركية",
  "kin_desc":"PFO · PSO · إيلوفيتش · ويبر-موريس · طاقة التنشيط أرينيوس",
  "c0_kin":"التركيز الابتدائي C₀ (mg/L)",
  "qt_max":"qₜ عند أقصى زمن","max_time":"أقصى زمن","max_qt":"أقصى qₜ",
  "kin_curves":"📈 منحنيات الحركية","wm_plot":"🌀 مخطط ويبر-موريس",
  "arrhenius_title":"🔥 طاقة التنشيط (أرينيوس)","arrhenius_desc":"أدخل ثوابت المعدل عند درجات حرارة مختلفة",
  "arrhenius_exp":"A = Ae × exp(−Ea/RT) — الميل = −Ea/R",
  "num_arr_temps":"عدد درجات الحرارة","arr_temp":"T (°C)","arr_k":"قيمة k",
  "ea_result":"طاقة التنشيط Eₐ","arr_r2":"R² أرينيوس",
  "ea_interp_phys":"أدمصاص فيزيائي (Eₐ < 40 kJ/mol)",
  "ea_interp_chem":"أدمصاص كيميائي (Eₐ > 40 kJ/mol)",
  "kin_min_pts":"📝 أدخل 4 نقاط على الأقل للبدء.",
  "thermo_header":"التحليل الثرموديناميكي","thermo_desc":"فانت هوف · ΔG° · ΔH° · ΔS°",
  "num_temps":"عدد درجات الحرارة","temp_pt":"T (°C)","kc_pt":"Kc / Kd",
  "vth_plot":"📈 مخطط فانت هوف","dg_plot":"📉 ΔG° مقابل الحرارة",
  "thermo_interp":"🔬 التفسير الثرموديناميكي",
  "enthalpy":"ΔH° (المحتوى الحراري)","entropy":"ΔS° (الإنتروبي)","dg_at_t1":"ΔG° عند T₁","vth_r2":"R²",
  "exo_title":"🔥 طارد للحرارة","endo_title":"❄️ ماص للحرارة",
  "exo_body":"تنبعث حرارة. الأدمصاص يتناقص بزيادة الحرارة.",
  "endo_body":"تمتص حرارة. الأدمصاص يتزايد بزيادة الحرارة.",
  "entropy_pos":"📈 اضطراب متزايد","entropy_neg":"📉 اضطراب متناقص",
  "entropy_pos_body":"ΔS° موجب — يزداد الاضطراب عند واجهة الصلب-المحلول.",
  "entropy_neg_body":"ΔS° سالب — واجهة أكثر انتظاماً.",
  "dg_table":"📋 ملخص ΔG°","spontaneous":"✅ عفوي عند جميع درجات الحرارة",
  "non_spontaneous":"⚠️ غير عفوي عند بعض درجات الحرارة",
  "exo_desc":"طارد (<0) / ماص (>0)","entropy_desc":"تغيّر الاضطراب","linear_fit":"جودة الانحدار",
  "duration":"مدة التجربة","peak_kin":"أقصى طاقة حركية",
  "valid_meas":"قياسات صحيحة","mean_across":"متوسط جميع النقاط",
  "peak_ads":"أقصى طاقة أدمصاص","max_eq_conc":"أقصى تركيز اتزاني",
  "report_header":"التقرير والتصدير","report_desc":"إنشاء تقرير Excel شامل بجميع النتائج",
  "exp_name":"اسم التجربة / رمز العينة","researcher":"اسم الباحث","notes_label":"ملاحظات",
  "export_includes":"📦 محتويات التصدير","generate_btn":"📊 إنشاء تقرير Excel",
  "download_btn":"⬇️ تحميل .xlsx","report_success":"✅ جاهز! اضغط للتحميل.",
  "session_results":"📊 نتائج الجلسة","no_results":"لا توجد نتائج — شغّل التحليل أولاً.",
  "iso_models_lbl":"📊 نماذج الأيزوثيرم","kin_models_lbl":"⏱️ النماذج الحركية","thermo_lbl":"🌡️ الثرموديناميكا",
  "demo_loaded":"تم تحميل البيانات التجريبية!","error_cols_iso":"يلزم عمودا: C0 و Absorbance",
  "error_cols_kin":"يلزم عمودا: Time و Absorbance",
  "conc_method":"طريقة التركيز: ",
  "interp_best":"أفضل نموذج مناسب هو","interp_r2":"بقيمة R² =",
  "interp_lang_fav":"أدمصاص مناسب (RL بين 0-1)",
  "interp_lang_unfav":"أدمصاص غير مناسب (RL > 1)",
  "interp_lang_irrev":"أدمصاص لا رجعي (RL = 0)",
  "interp_frnd_good":"أدمصاص غير متجانس مناسب (1/n < 1)",
  "interp_frnd_poor":"ظروف أدمصاص ضعيفة (1/n > 1)",
  "data_quality":"📊 جودة البيانات",
  "dq_cv":"معامل التباين","dq_range":"نطاق qₑ","dq_pts":"نقاط صحيحة",
  "dq_mono":"رتيب","dq_yes":"نعم ✅","dq_no":"لا ⚠️",
}}

def t(k): return TRANS[st.session_state.lang].get(k, k)

is_ar   = st.session_state.lang == "ar"
is_dark = st.session_state.dark
fd      = "rtl" if is_ar else "ltr"
ff      = "Cairo,Inter,sans-serif" if is_ar else "Inter,sans-serif"

# ═══════════════════════════════════════════════════════════════════════════════
#  THEME TOKENS
# ═══════════════════════════════════════════════════════════════════════════════
if is_dark:
    BG        = "#070b14"; SURF  = "#0d1628"; CARD  = "rgba(13,22,40,0.9)"
    CARD2     = "rgba(30,41,59,0.45)"; BR    = "rgba(99,102,241,0.22)"; BRS = "rgba(99,102,241,0.5)"
    TX        = "#e2e8f0"; TXS   = "#94a3b8"; TXM   = "rgba(148,163,184,0.5)"
    AC        = "#818cf8"; AC2   = "#c084fc"; ACB   = "rgba(99,102,241,0.1)"
    AG        = "linear-gradient(135deg,#6366f1,#8b5cf6)"
    INP       = "rgba(7,11,20,0.85)"; TABB = "rgba(7,11,20,0.7)"; SBG = "linear-gradient(180deg,#0d1628,#070b14)"
    PP        = "rgba(7,11,20,0)"; PG = "rgba(99,102,241,0.09)"; PT = "rgba(99,102,241,0.25)"
    SUCBG     = "rgba(8,35,18,0.7)"; SUCBR = "rgba(34,197,94,0.35)"
    WARNBG    = "rgba(40,18,8,0.7)"; WARNBR = "rgba(251,146,60,0.35)"
    SCROLLC   = "rgba(99,102,241,0.4)"
else:
    BG        = "#eef2ff"; SURF  = "#ffffff"; CARD  = "rgba(255,255,255,0.97)"
    CARD2     = "rgba(238,242,255,0.8)"; BR    = "rgba(99,102,241,0.16)"; BRS = "rgba(99,102,241,0.45)"
    TX        = "#1e293b"; TXS   = "#475569"; TXM   = "rgba(71,85,105,0.55)"
    AC        = "#4f46e5"; AC2   = "#7c3aed"; ACB   = "rgba(99,102,241,0.07)"
    AG        = "linear-gradient(135deg,#4f46e5,#7c3aed)"
    INP       = "#ffffff"; TABB = "rgba(238,242,255,0.9)"; SBG = "linear-gradient(180deg,#ffffff,#eef2ff)"
    PP        = "rgba(238,242,255,0)"; PG = "rgba(99,102,241,0.07)"; PT = "rgba(99,102,241,0.2)"
    SUCBG     = "rgba(220,252,231,0.9)"; SUCBR = "rgba(22,163,74,0.3)"
    WARNBG    = "rgba(255,237,213,0.9)"; WARNBR = "rgba(234,88,12,0.3)"
    SCROLLC   = "rgba(99,102,241,0.3)"

PALETTE = ["#6366f1","#c084fc","#34d399","#fb923c","#22d3ee","#f472b6","#fbbf24","#a3e635"]
PTBASE = dict(paper_bgcolor=PP,plot_bgcolor=PP,
              font=dict(family=ff,color=TXS,size=11),
              xaxis=dict(gridcolor=PG,linecolor=PT,tickcolor=PT,
                         title_font=dict(color=TXS,size=12),tickfont=dict(color=TXS)),
              yaxis=dict(gridcolor=PG,linecolor=PT,tickcolor=PT,
                         title_font=dict(color=TXS,size=12),tickfont=dict(color=TXS)),
              legend=dict(bgcolor=CARD,bordercolor=BR,borderwidth=1,font=dict(color=TXS,size=10)),
              hoverlabel=dict(bgcolor=SURF,bordercolor=BRS,font=dict(color=TX,size=11)),
              margin=dict(l=50,r=30,t=40,b=50))

# ═══════════════════════════════════════════════════════════════════════════════
#  GLOBAL CSS
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Cairo:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
html,body,[class*="css"],.stApp,.main{{font-family:{ff}!important;direction:{fd};}}
.stApp{{background:{BG};min-height:100vh;}}
.main .block-container{{padding:1.4rem 1.8rem 3rem;max-width:1450px;}}
[data-testid="stSidebar"]{{background:{SBG}!important;border-{"left" if is_ar else "right"}:1px solid {BR};}}
[data-testid="stSidebar"] .block-container{{padding-top:.7rem!important;}}
[data-baseweb="tab-list"]{{background:{TABB}!important;border-radius:14px!important;padding:4px!important;border:1px solid {BR}!important;gap:2px!important;}}
[data-baseweb="tab"]{{border-radius:10px!important;color:{TXS}!important;font-weight:500!important;font-size:.81rem!important;padding:.42rem .95rem!important;transition:all .2s!important;font-family:{ff}!important;}}
[aria-selected="true"][data-baseweb="tab"]{{background:{AG}!important;color:white!important;font-weight:700!important;}}
[data-testid="stNumberInput"] input,[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{{background:{INP}!important;border:1px solid {BR}!important;border-radius:10px!important;color:{TX}!important;font-family:{ff}!important;direction:ltr;}}
[data-testid="stNumberInput"] input:focus,[data-testid="stTextInput"] input:focus{{border-color:{BRS}!important;box-shadow:0 0 0 3px {ACB}!important;}}
[data-testid="stSelectbox"]>div>div{{background:{INP}!important;border:1px solid {BR}!important;border-radius:10px!important;color:{TX}!important;}}
.stButton>button{{background:{AG}!important;color:white!important;border:none!important;border-radius:10px!important;font-weight:600!important;font-size:.83rem!important;transition:all .25s!important;font-family:{ff}!important;}}
.stButton>button:hover{{transform:translateY(-1px)!important;filter:brightness(1.1)!important;box-shadow:0 8px 22px {ACB}!important;}}
[data-testid="stDownloadButton"]>button{{background:linear-gradient(135deg,#059669,#10b981)!important;color:white!important;border:none!important;border-radius:10px!important;font-weight:600!important;font-family:{ff}!important;}}
[data-testid="stDataFrame"]{{border-radius:12px!important;overflow:hidden!important;border:1px solid {BR}!important;}}
[data-testid="stRadio"] label,[data-testid="stToggle"] label{{color:{TXS}!important;font-size:.83rem!important;font-family:{ff}!important;}}
[data-testid="stMultiSelect"]>div>div{{background:{INP}!important;border:1px solid {BR}!important;border-radius:10px!important;}}
[data-baseweb="tag"]{{background:{ACB}!important;border:1px solid {BR}!important;border-radius:6px!important;}}
details{{background:{CARD}!important;border:1px solid {BR}!important;border-radius:10px!important;}}
details summary{{color:{TXS}!important;font-family:{ff}!important;font-size:.83rem!important;}}
[data-testid="stSlider"] label{{color:{TXS}!important;font-size:.82rem!important;font-family:{ff}!important;}}
[data-testid="stNumberInput"] label,[data-testid="stTextInput"] label,
[data-testid="stTextArea"] label,[data-testid="stSelectbox"] label,
[data-testid="stMultiSelect"] label{{color:{TXS}!important;font-size:.81rem!important;font-weight:500!important;font-family:{ff}!important;}}
::-webkit-scrollbar{{width:4px;height:4px;}}::-webkit-scrollbar-track{{background:transparent;}}
::-webkit-scrollbar-thumb{{background:{SCROLLC};border-radius:10px;}}

/* ── components ── */
.ada-metric{{background:{CARD};border:1px solid {BR};border-radius:14px;padding:1rem 1.15rem;text-align:center;position:relative;overflow:hidden;}}
.ada-metric::after{{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:{AG};border-radius:14px 14px 0 0;}}
.ada-metric-val{{font-size:1.5rem;font-weight:700;color:{AC};font-family:'JetBrains Mono',monospace;line-height:1.1;}}
.ada-metric-lbl{{font-size:.67rem;font-weight:600;color:{TXM};letter-spacing:.8px;text-transform:uppercase;margin-top:.25rem;font-family:{ff};}}
.ada-metric-sub{{font-size:.69rem;color:{TXM};margin-top:.12rem;font-family:{ff};}}
.ada-section{{display:flex;align-items:center;gap:.7rem;margin-bottom:1.1rem;padding-bottom:.7rem;border-bottom:1px solid {BR};direction:{fd};}}
.ada-icon{{width:36px;height:36px;background:{AG};border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0;}}
.ada-title{{font-size:1.08rem;font-weight:700;color:{TX};font-family:{ff};}}
.ada-desc{{font-size:.76rem;color:{TXM};font-family:{ff};}}
.ada-eq{{background:{ACB};border:1px solid {BR};border-{"right" if is_ar else "left"}:3px solid {AC};border-radius:8px;padding:.8rem 1.05rem;font-size:.79rem;color:{TXS};font-family:{ff};direction:{fd};line-height:1.7;}}
.ada-card{{background:{CARD};border:1px solid {BR};border-radius:12px;padding:.8rem 1.05rem;margin-bottom:.5rem;direction:{fd};}}
.ada-card.best{{border-color:{SUCBR};background:{SUCBG};}}
.ada-card.exo{{border-color:{WARNBR};background:{WARNBG};}}
.ada-badge{{background:{AG};color:white;font-size:.58rem;font-weight:700;padding:.14rem .48rem;border-radius:20px;letter-spacing:.5px;font-family:{ff};}}
.param-row{{display:flex;justify-content:space-between;align-items:center;padding:.28rem 0;border-bottom:1px solid {BR};direction:{fd};}}
.param-key{{color:{TXM};font-size:.78rem;font-family:{ff};}}
.param-val{{color:{AC};font-family:'JetBrains Mono',monospace;font-size:.79rem;font-weight:600;}}
.sb-section{{font-size:.61rem;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:{AC};opacity:.7;padding:.65rem .2rem .28rem;font-family:{ff};direction:{fd};}}
.ada-hero{{background:{CARD};border:1px solid {BR};border-radius:18px;padding:1.5rem 2rem;margin-bottom:1.3rem;text-align:center;position:relative;overflow:hidden;direction:{fd};}}
.ada-hero::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:{AG};border-radius:18px 18px 0 0;}}
.ada-hero-title{{font-size:1.75rem;font-weight:800;background:{AG};-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;font-family:{ff};margin:0;}}
.ada-hero-sub{{color:{TXS};font-size:.83rem;margin-top:.3rem;font-family:{ff};}}
.ada-hero-badge{{display:inline-block;background:{AG};color:white;font-size:.61rem;font-weight:700;padding:.18rem .65rem;border-radius:20px;letter-spacing:.8px;margin-top:.5rem;font-family:{ff};}}
.feat-card{{background:{CARD};border:1px solid {BR};border-radius:14px;padding:1.1rem 1.2rem;height:100%;transition:all .25s;direction:{fd};}}
.feat-card:hover{{border-color:{BRS};transform:translateY(-2px);box-shadow:0 8px 24px {ACB};}}
.feat-icon{{font-size:1.6rem;margin-bottom:.5rem;}}
.feat-title{{font-size:.88rem;font-weight:700;color:{TX};font-family:{ff};}}
.feat-desc{{font-size:.75rem;color:{TXM};margin-top:.2rem;font-family:{ff};line-height:1.5;}}
.guide-step{{display:flex;gap:.7rem;align-items:flex-start;padding:.65rem .9rem;background:{CARD};border:1px solid {BR};border-radius:10px;margin-bottom:.4rem;direction:{fd};}}
.guide-num{{width:22px;height:22px;background:{AG};border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:.68rem;font-weight:700;color:white;flex-shrink:0;font-family:{ff};}}
.guide-text{{font-size:.8rem;color:{TXS};font-family:{ff};line-height:1.55;}}
.status-chip{{display:inline-flex;align-items:center;gap:.35rem;padding:.28rem .65rem;border-radius:20px;font-size:.72rem;font-weight:600;font-family:{ff};}}
.status-chip.done{{background:{SUCBG};color:#34d399;border:1px solid {SUCBR};}}
.status-chip.pending{{background:{ACB};color:{TXM};border:1px solid {BR};}}
.interp-box{{background:{ACB};border:1px solid {BRS};border-radius:12px;padding:1rem 1.2rem;direction:{fd};}}
.interp-title{{font-size:.85rem;font-weight:700;color:{AC};font-family:{ff};margin-bottom:.5rem;}}
.interp-line{{font-size:.79rem;color:{TXS};font-family:{ff};line-height:1.7;padding:.15rem 0;border-bottom:1px solid {BR};}}
.model-ref-row{{display:flex;justify-content:space-between;padding:.32rem 0;border-bottom:1px solid {BR};direction:{fd};}}
.ada-footer{{text-align:center;color:{TXM};font-size:.69rem;padding:1.6rem 0 .4rem;border-top:1px solid {BR};margin-top:2rem;font-family:{ff};direction:{fd};}}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  MATH MODELS
# ═══════════════════════════════════════════════════════════════════════════════
def langmuir(Ce,qm,KL):   return (qm*KL*Ce)/(1+KL*Ce)
def freundlich(Ce,KF,n):  return KF*(Ce**(1/n))
def temkin(Ce,AT,B):      return B*np.log(AT*Ce)
def make_dr(T_K, mw_gpmol):
    """Factory: returns D-R model with fixed T(K) and M(g/mol).
    ε = RT·ln(1+1/Ce_mol) [kJ/mol], β fitted [mol²/kJ²], E=1/√(2β) [kJ/mol]"""
    R_kJ = 8.314e-3  # kJ/(mol·K)
    def _dr(Ce, qm, beta):
        Ce_mol = np.maximum(Ce / (mw_gpmol * 1000.0), 1e-15)  # mg/L → mol/L
        eps = R_kJ * T_K * np.log(1.0 + 1.0 / Ce_mol)         # kJ/mol
        return qm * np.exp(-beta * eps**2)
    return _dr
def sips(Ce,qm,Ks,ns):    return (qm*Ks*Ce**ns)/(1+Ks*Ce**ns)
def bet_model(Ce,Cs,qm_b,C_b):
    x=Ce/Cs; return (C_b*qm_b*x)/((1-x)*(1+(C_b-1)*x+1e-10))
def pfo(t,qe,k1):         return qe*(1-np.exp(-k1*t))
def pso(t,qe,k2):         return (k2*qe**2*t)/(1+k2*qe*t)
def elovich(t,a,b):       return (1/b)*np.log(a*b*t+1)
def wm(t,kid,C):          return kid*np.sqrt(t)+C

def r2s(y,yp):
    ss=np.sum((y-yp)**2); st=np.sum((y-np.mean(y))**2)
    return float(1-ss/st) if st!=0 else 0.0
def rmse(y,yp): return float(np.sqrt(np.mean((y-yp)**2)))
def chi2(y,yp): return float(np.sum((y-yp)**2/(np.abs(yp)+1e-10)))

def do_fit(func,x,y,p0,bounds=(0,np.inf)):
    try:
        po,pcov=curve_fit(func,x,y,p0=p0,bounds=bounds,maxfev=15000)
        yp=func(x,*po); return po,pcov,r2s(y,yp),rmse(y,yp),chi2(y,yp)
    except: return None,None,0.,999.,999.

def get_c(a,method,**kw):
    if method=="beer": return (a/(kw['epsilon']*kw['path_length']))*kw['mw']*1000
    return (a-kw['intercept'])/kw['slope']

# ═══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="text-align:center;padding:.7rem 0 .4rem;">
        <div style="font-size:1.8rem;margin-bottom:.2rem;">🔬</div>
        <div style="font-size:1.02rem;font-weight:800;background:{AG};
            -webkit-background-clip:text;-webkit-text-fill-color:transparent;
            background-clip:text;font-family:{ff};">{t("app_title")}</div>
        <div style="font-size:.58rem;color:{TXM};letter-spacing:1px;margin-top:.12rem;font-family:{ff};">
            SCIENTIFIC ANALYSIS SUITE</div>
        <span style="background:{AG};color:white;font-size:.56rem;font-weight:700;
            padding:.13rem .48rem;border-radius:20px;letter-spacing:.5px;font-family:{ff};">v2.0 PRO</span>
    </div>
    <hr style="border:none;border-top:1px solid {BR};margin:.6rem 0;">
    """, unsafe_allow_html=True)

    c1,c2=st.columns(2)
    with c1:
        if st.button(t("lang_btn"),use_container_width=True,key="lang_t"):
            st.session_state.lang="ar" if st.session_state.lang=="en" else "en"; st.rerun()
    with c2:
        if st.button(t("dark_btn") if is_dark else t("light_btn"),use_container_width=True,key="thm_t"):
            st.session_state.dark=not st.session_state.dark; st.rerun()

    st.markdown(f'<div class="sb-section">{t("cal_settings")}</div>',unsafe_allow_html=True)
    cal_opts=[t("cal_beer"),t("cal_curve")]
    cal_method=st.selectbox(t("cal_method_label"),cal_opts,label_visibility="collapsed")
    if cal_method==t("cal_beer"):
        eps_v=st.number_input(t("epsilon"),value=15000.0,format="%.1f")
        pl_v=st.number_input(t("path_len"),value=1.00,format="%.2f")
        mw_v=st.number_input(t("molar_mass"),value=319.85,format="%.2f")
        cal_kw=dict(epsilon=eps_v,path_length=pl_v,mw=mw_v); cal_key="beer"
    else:
        sl_v=st.number_input(t("slope"),value=0.1000,format="%.4f")
        ic_v=st.number_input(t("intercept"),value=0.0000,format="%.4f")
        cal_kw=dict(slope=sl_v,intercept=ic_v); cal_key="curve"

    st.markdown(f'<div class="sb-section">{t("visualization")}</div>',unsafe_allow_html=True)
    show_res=st.toggle(t("show_residuals"),value=False)
    mk_sz=st.slider(t("marker_size"),5,16,9)

    st.markdown(f'<div class="sb-section">{t("quick_actions")}</div>',unsafe_allow_html=True)
    if st.button(t("demo_iso"),use_container_width=True):
        st.session_state.iso_data=DEMO_ISO.copy(); st.toast(t("demo_loaded"),icon="🧪")
    if st.button(t("demo_kin"),use_container_width=True):
        st.session_state.kin_data=DEMO_KIN.copy(); st.toast(t("demo_loaded"),icon="⏱")

    st.markdown(f"""
    <hr style="border:none;border-top:1px solid {BR};margin:.8rem 0;">
    <div style="font-size:.61rem;color:{TXM};text-align:center;line-height:1.9;font-family:{ff};">
        Langmuir · Freundlich · Temkin · D-R · Sips · BET<br>
        PFO · PSO · Elovich · Weber-Morris · Arrhenius
    </div>
    <hr style="border:none;border-top:1px solid {BR};margin:.6rem 0;">
    <div style="text-align:center;font-size:.58rem;color:{AC};font-weight:600;font-family:{ff};letter-spacing:.3px;line-height:1.7;padding-bottom:.3rem;">
        Diana Raie<br>Abdelbaset A. A. Diyab
    </div>""",unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  HERO
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="ada-hero">
    <p class="ada-hero-title">🔬 {t("app_title")}</p>
    <p class="ada-hero-sub">{t("app_subtitle")}</p>
    <span class="ada-hero-badge">{t("app_badge")}</span>
</div>""",unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  TABS
# ═══════════════════════════════════════════════════════════════════════════════
tab_home,tab_cal,tab_iso,tab_kin,tab_thermo,tab_rep = st.tabs([
    t("tab_home"),t("tab_cal"),t("tab_iso"),t("tab_kin"),t("tab_thermo"),t("tab_report")])

# ─────────────────────────────────────────────────────────────────────────────
#  HOME
# ─────────────────────────────────────────────────────────────────────────────
with tab_home:
    st.markdown(f"""
    <div class="ada-section"><div class="ada-icon">🏠</div>
    <div><div class="ada-title">{t("home_welcome")}</div>
    <div class="ada-desc">{t("home_sub")}</div></div></div>""",unsafe_allow_html=True)

    # Feature cards
    st.markdown(f"<div style='font-size:.95rem;font-weight:700;color:{TX};font-family:{ff};margin-bottom:.8rem;'>{t('home_feat_title')}</div>",unsafe_allow_html=True)
    feats=[
        ("🧪",t("feat1_t"),t("feat1_d")),("⏱️",t("feat2_t"),t("feat2_d")),
        ("🌡️",t("feat3_t"),t("feat3_d")),("📊",t("feat4_t"),t("feat4_d")),
        ("📐",t("feat5_t"),t("feat5_d")),("📄",t("feat6_t"),t("feat6_d")),
    ]
    cols=st.columns(3,gap="medium")
    for i,(ic,ti,de) in enumerate(feats):
        with cols[i%3]:
            st.markdown(f"""
            <div class="feat-card">
                <div class="feat-icon">{ic}</div>
                <div class="feat-title">{ti}</div>
                <div class="feat-desc">{de}</div>
            </div><br>""",unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    lft,rgt=st.columns([3,2],gap="large")

    with lft:
        st.markdown(f"<div style='font-size:.95rem;font-weight:700;color:{TX};font-family:{ff};margin-bottom:.7rem;'>{t('guide_title')}</div>",unsafe_allow_html=True)
        for i,step in enumerate([t(f"guide{j}") for j in range(1,6)],1):
            st.markdown(f"""
            <div class="guide-step">
                <div class="guide-num">{i}</div>
                <div class="guide-text">{step}</div>
            </div>""",unsafe_allow_html=True)

        st.markdown(f"<br><div style='font-size:.95rem;font-weight:700;color:{TX};font-family:{ff};margin-bottom:.7rem;'>{t('model_ref')}</div>",unsafe_allow_html=True)
        st.markdown(f"""
        <div class="ada-card">
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:.3rem;">
                <div>
                    <div style="color:{AC};font-size:.73rem;font-weight:700;margin-bottom:.3rem;font-family:{ff};">Isotherm Models</div>
                    {"".join([f'<div class="model-ref-row"><span style="color:{TXS};font-size:.76rem;font-family:{ff};">{n}</span><span style="color:{TXM};font-size:.7rem;font-family:{ff};">{d}</span></div>'
                     for n,d in [("Langmuir","Homogeneous, monolayer"),("Freundlich","Heterogeneous, multilayer"),
                                  ("Temkin","Heat effect, linear"),("D-R","Adsorption energy"),
                                  ("Sips","Combined L+F"),("BET","Multilayer, porous")]])}
                </div>
                <div>
                    <div style="color:{AC};font-size:.73rem;font-weight:700;margin-bottom:.3rem;font-family:{ff};">Kinetic Models</div>
                    {"".join([f'<div class="model-ref-row"><span style="color:{TXS};font-size:.76rem;font-family:{ff};">{n}</span><span style="color:{TXM};font-size:.7rem;font-family:{ff};">{d}</span></div>'
                     for n,d in [("PFO","Physical diffusion"),("PSO","Chemical interaction"),
                                  ("Elovich","Chemisorption"),("Weber-Morris","Intraparticle"),
                                  ("Arrhenius","Activation energy")]])}
                </div>
            </div>
        </div>""",unsafe_allow_html=True)

    with rgt:
        st.markdown(f"<div style='font-size:.95rem;font-weight:700;color:{TX};font-family:{ff};margin-bottom:.7rem;'>{t('session_title')}</div>",unsafe_allow_html=True)
        for key,lbl in [("iso_results",t("iso_status")),("kin_results",t("kin_status")),("thermo_results",t("thermo_status"))]:
            done=key in st.session_state
            cls="done" if done else "pending"
            txt=t("done") if done else t("not_run")
            st.markdown(f"""
            <div class="ada-card" style="display:flex;justify-content:space-between;align-items:center;margin-bottom:.4rem;">
                <span style="color:{TX};font-size:.83rem;font-weight:600;font-family:{ff};">{lbl}</span>
                <span class="status-chip {cls}">{txt}</span>
            </div>""",unsafe_allow_html=True)

        # Mini summary if results exist
        if "iso_results" in st.session_state:
            best=st.session_state.get("iso_best","")
            res=st.session_state["iso_results"].get(best,{})
            r2v=res.get("r2",0)
            st.markdown(f"""
            <div class="interp-box" style="margin-top:.8rem;">
                <div class="interp-title">📊 Best Isotherm</div>
                <div class="interp-line">🥇 {best} — R² = {r2v:.4f}</div>
            </div>""",unsafe_allow_html=True)

        if "kin_results" in st.session_state:
            bestk=st.session_state.get("kin_best","")
            resk=st.session_state["kin_results"].get(bestk,{})
            r2k=resk.get("r2",0)
            st.markdown(f"""
            <div class="interp-box" style="margin-top:.6rem;">
                <div class="interp-title">⏱️ Best Kinetic</div>
                <div class="interp-line">🥇 {bestk} — R² = {r2k:.4f}</div>
            </div>""",unsafe_allow_html=True)

        if "thermo_results" in st.session_state:
            tr=st.session_state["thermo_results"]
            st.markdown(f"""
            <div class="interp-box" style="margin-top:.6rem;">
                <div class="interp-title">🌡️ Thermodynamics</div>
                <div class="interp-line">ΔH° = {tr['delta_H']:.3f} kJ/mol</div>
                <div class="interp-line">ΔS° = {tr['delta_S']:.3f} J/mol·K</div>
            </div>""",unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  CALIBRATION
# ─────────────────────────────────────────────────────────────────────────────
with tab_cal:
    st.markdown(f"""
    <div class="ada-section"><div class="ada-icon">⚗️</div>
    <div><div class="ada-title">{t("cal_header")}</div>
    <div class="ada-desc">{t("cal_desc")}</div></div></div>""",unsafe_allow_html=True)

    lc,rc=st.columns(2,gap="large")
    with lc:
        st.markdown(f'<div class="ada-eq">📐 {t("cal_beer") if cal_key=="beer" else t("cal_curve")}<br>{"A = ε · b · C → C(mg/L) = C(mol/L) × M × 1000" if cal_key=="beer" else "C = (A − intercept) / slope"}</div>',unsafe_allow_html=True)
        st.markdown(f"<br><b style='color:{TXS};font-size:.88rem;font-family:{ff};'>{t('live_preview')}</b>",unsafe_allow_html=True)
        ta=st.number_input(t("test_abs"),value=0.500,format="%.3f",step=.001)
        cv=get_c(np.array([ta]),cal_key,**cal_kw)[0]
        st.markdown(f"""
        <div class="ada-card" style="margin-top:.55rem;">
            <div style="color:{TXM};font-size:.68rem;letter-spacing:.8px;text-transform:uppercase;font-family:{ff};">{t("calc_conc")}</div>
            <div style="font-size:1.9rem;font-weight:700;color:{AC};font-family:'JetBrains Mono',monospace;margin-top:.18rem;">
                {cv:.4f} <span style="font-size:.85rem;color:{TXM};">mg/L</span>
            </div>
            <div style="font-size:.7rem;color:{TXM};margin-top:.12rem;font-family:{ff};">A = {ta:.3f}</div>
        </div>""",unsafe_allow_html=True)

    with rc:
        st.markdown(f"<b style='color:{TXS};font-size:.88rem;font-family:{ff};'>{t('cal_curve_vis')}</b>",unsafe_allow_html=True)
        ar=np.linspace(.01,1.6,300)
        cr=get_c(ar,cal_key,**cal_kw)
        fg=go.Figure(); fg.add_trace(go.Scatter(x=cr,y=ar,mode='lines',
            line=dict(color=AC,width=2.5),fill='tozeroy',fillcolor=ACB))
        fg.update_layout(**PTBASE,height=270,xaxis_title=t("conc_mgl"),yaxis_title=t("absorbance"),showlegend=False)
        st.plotly_chart(fg,use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
#  ISOTHERMS
# ─────────────────────────────────────────────────────────────────────────────
with tab_iso:
    st.markdown(f"""
    <div class="ada-section"><div class="ada-icon">📊</div>
    <div><div class="ada-title">{t("iso_header")}</div>
    <div class="ada-desc">{t("iso_desc")}</div></div></div>""",unsafe_allow_html=True)

    ic1,ic2,ic3=st.columns(3)
    with ic1: v_l=st.number_input(t("vol"),value=.050,format="%.3f",key="vl")
    with ic2: m_g=st.number_input(t("mass"),value=.100,format="%.3f",key="mg")
    with ic3: iso_sel=st.multiselect(t("models_fit"),
        ["Langmuir","Freundlich","Temkin","D-R","Sips","BET"],
        default=["Langmuir","Freundlich","Temkin","D-R","Sips"],key="isel")

    if "D-R" in iso_sel or "Temkin" in iso_sel:
        dr1,dr2=st.columns(2)
        with dr1:
            T_iso_C=st.number_input("Temperature T (°C) — for D-R & Temkin",value=25.0,format="%.1f",key="t_iso")
        with dr2:
            mw_default=float(cal_kw.get("mw",100.0))
            mw_dr=st.number_input("Molar Mass M (g/mol) — for D-R",value=mw_default,format="%.2f",key="mw_dr")
        T_K_iso=T_iso_C+273.15
    else:
        T_K_iso=298.15; mw_dr=cal_kw.get("mw",100.0)

    st.markdown(f"<hr style='border:none;border-top:1px solid {BR};margin:.5rem 0;'>",unsafe_allow_html=True)
    md=st.radio(t("data_input"),[t("upload"),t("manual")],horizontal=True,key="miso")

    if md==t("upload"):
        uf=st.file_uploader(t("upload_hint_iso"),type=["xlsx","csv"],key="iuf")
        if uf:
            try:
                dr=pd.read_excel(uf) if uf.name.endswith(".xlsx") else pd.read_csv(uf)
                if "C0" in dr.columns and "Absorbance" in dr.columns:
                    st.session_state.iso_data=dr[["C0","Absorbance"]].copy()
                else: st.error(t("error_cols_iso"))
            except Exception as e: st.error(str(e))
    else:
        st.session_state.iso_data=st.data_editor(st.session_state.iso_data,num_rows="dynamic",
            use_container_width=True,key="ied",
            column_config={"C0":st.column_config.NumberColumn("C₀ (mg/L)",format="%.3f"),
                           "Absorbance":st.column_config.NumberColumn("Absorbance",format="%.4f")})

    df_i=st.session_state.iso_data.copy()
    if df_i is not None and len(df_i)>=3:
        try:
            df_i["Ce"]=get_c(df_i["Absorbance"].values,cal_key,**cal_kw)
            df_i["qe"]=((df_i["C0"]-df_i["Ce"])*v_l)/m_g
            df_i["Rem%"]=((df_i["C0"]-df_i["Ce"])/df_i["C0"])*100
            df_i=df_i[df_i["Ce"]>0].reset_index(drop=True)
            Ce,qe=df_i["Ce"].values,df_i["qe"].values

            # Data quality
            cv_val=float(np.std(qe)/np.mean(qe)*100) if np.mean(qe)!=0 else 0
            mono=bool(np.all(np.diff(qe)>=0) or np.all(np.diff(qe)<=0))

            k1_,k2_,k3_,k4_=st.columns(4)
            for col,(val,lbl,sub) in zip([k1_,k2_,k3_,k4_],[
                (f"{df_i['Rem%'].mean():.1f}%",t("avg_removal"),t("mean_across")),
                (f"{qe.max():.2f}",t("max_qe"),t("peak_ads")),
                (f"{Ce.max():.2f}",t("max_ce"),t("max_eq_conc")),
                (f"{len(df_i)}",t("data_pts"),t("valid_meas"))]):
                with col:
                    st.markdown(f"""<div class="ada-metric"><div class="ada-metric-val">{val}</div>
                        <div class="ada-metric-lbl">{lbl}</div><div class="ada-metric-sub">{sub}</div></div>""",unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)

            # Fit models — D-R uses factory with T and MW for correct ε = RT·ln(1+1/Ce_mol)
            _dr_fn=make_dr(T_K_iso, float(mw_dr))
            funcs={"Langmuir":langmuir,"Freundlich":freundlich,"Temkin":temkin,"D-R":_dr_fn,"Sips":sips,"BET":bet_model}
            # D-R: fit (qm, β) where β [mol²/kJ²]; β₀ ≈ 1/(2·E₀²), E₀≈5 kJ/mol → β₀≈0.02
            p0s={"Langmuir":[qe.max(),.1],"Freundlich":[1.,2.],"Temkin":[1.,qe.max()/5],
                 "D-R":[qe.max(),0.02],"Sips":[qe.max(),.1,1.0],"BET":[Ce.max()*2,qe.max(),5.]}
            bds={"Langmuir":(0,np.inf),"Freundlich":(0,np.inf),"Temkin":([1e-10,1e-10],[np.inf,np.inf]),
                 "D-R":([0,1e-12],[np.inf,np.inf]),"Sips":([0,0,.1],[np.inf,np.inf,5.]),
                 "BET":([Ce.max()*1.01,0,0],[np.inf,np.inf,np.inf])}

            iso_res={}
            for nm in iso_sel:
                if nm in("Temkin","D-R","Sips") and Ce.min()<=0: continue
                po,pcov_,rv,rm,rc_=do_fit(funcs[nm],Ce,qe,p0s[nm],bds[nm])
                if po is not None:
                    if nm=="Langmuir":
                        RL=1/(1+po[1]*Ce.max())
                        pms={"qₘₐₓ (mg/g)":po[0],"KL (L/mg)":po[1],"RL":RL}
                    elif nm=="Freundlich":
                        pms={"KF":po[0],"1/n":1/po[1],"n":po[1]}
                    elif nm=="Temkin":
                        R_J=8.314; bT=R_J*T_K_iso/po[1] if po[1]>0 else float("inf")
                        pms={"AT (L/g)":po[0],"B (mg/g)":po[1],"bT (J/mol)":bT}
                    elif nm=="D-R":
                        beta_v=po[1]
                        E_kJ=1.0/math.sqrt(2.0*beta_v) if beta_v>0 else 0.0
                        if E_kJ<8:    ads_type="Physisorption"
                        elif E_kJ<16: ads_type="Ion Exchange"
                        else:         ads_type="Chemisorption"
                        pms={"qₘ (mg/g)":po[0],"β (mol²/kJ²)":beta_v,"E (kJ/mol)":E_kJ,"Type":ads_type}
                    elif nm=="Sips":
                        pms={"qₘₐₓ (mg/g)":po[0],"Ks":po[1],"ns":po[2]}
                    else:  # BET
                        pms={"Cs (mg/L)":po[0],"qₘ (mg/g)":po[1],"C_BET":po[2]}
                    iso_res[nm]={"popt":po,"pcov":pcov_,"r2":rv,"rmse":rm,"chi2":rc_,"params":pms}

            if iso_res:
                best_iso=max(iso_res,key=lambda k:iso_res[k]["r2"])

                ch_col,st_col=st.columns([3,2],gap="large")

                with ch_col:
                    # Main isotherm plot
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('iso_curves')}</b>",unsafe_allow_html=True)
                    fig_i=go.Figure()
                    fig_i.add_trace(go.Scatter(x=Ce,y=qe,mode="markers",name="Exp.",
                        marker=dict(color="#f87171",size=mk_sz,symbol="circle",
                                    line=dict(color="rgba(248,113,113,0.3)",width=1))))
                    CeL=np.linspace(Ce.min()*.4,Ce.max()*1.25,400)
                    for i,(nm,res) in enumerate(iso_res.items()):
                        try:
                            yL=funcs[nm](CeL,*res["popt"])
                            yL=np.where(np.isfinite(yL),yL,np.nan)
                            is_b=(nm==best_iso)
                            fig_i.add_trace(go.Scatter(x=CeL,y=yL,mode="lines",
                                name=f"{nm} R²={res['r2']:.4f}"+(" ★" if is_b else ""),
                                line=dict(color=PALETTE[i],width=3 if is_b else 1.8,
                                          dash="solid" if is_b else "dot")))
                        except: pass
                    fig_i.update_layout(**PTBASE,height=360,
                        xaxis_title="Cₑ (mg/L)",yaxis_title="qₑ (mg/g)")
                    st.plotly_chart(fig_i,use_container_width=True)

                    # R² comparison bar chart
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('r2_chart')}</b>",unsafe_allow_html=True)
                    names_r=[nm for nm in iso_res]; r2_vals=[iso_res[nm]["r2"] for nm in names_r]
                    rmse_vals=[iso_res[nm]["rmse"] for nm in names_r]
                    fig_bar=make_subplots(rows=1,cols=2,subplot_titles=["R²","RMSE"])
                    fig_bar.add_trace(go.Bar(x=names_r,y=r2_vals,marker_color=PALETTE[:len(names_r)],
                        name="R²",text=[f"{v:.4f}" for v in r2_vals],textposition="outside"),row=1,col=1)
                    fig_bar.add_trace(go.Bar(x=names_r,y=rmse_vals,
                        marker_color=[PALETTE[i] for i in range(len(names_r))],
                        name="RMSE",text=[f"{v:.3f}" for v in rmse_vals],textposition="outside"),row=1,col=2)
                    fig_bar.update_layout(**PTBASE,height=230,showlegend=False)
                    fig_bar.update_yaxes(gridcolor=PG)
                    fig_bar.update_annotations(font_color=TXS)
                    st.plotly_chart(fig_bar,use_container_width=True)

                    # Linear transform plots
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('linear_plots')}</b>",unsafe_allow_html=True)
                    lt1,lt2=st.columns(2,gap="small")
                    with lt1:
                        # Langmuir linear: Ce/qe vs Ce
                        Ce_qe=Ce/qe; sl_l,ic_l=np.polyfit(Ce,Ce_qe,1)
                        r2_l_lin=r2s(Ce_qe,sl_l*Ce+ic_l)
                        fig_ll=go.Figure()
                        fig_ll.add_trace(go.Scatter(x=Ce,y=Ce_qe,mode="markers",
                            marker=dict(color=PALETTE[0],size=mk_sz)))
                        xln=np.linspace(Ce.min()*.8,Ce.max()*1.1,100)
                        fig_ll.add_trace(go.Scatter(x=xln,y=sl_l*xln+ic_l,mode="lines",
                            line=dict(color=PALETTE[0],width=2),name=f"R²={r2_l_lin:.4f}"))
                        fig_ll.update_layout(**PTBASE,height=220,
                            title=dict(text=t("langmuir_linear"),font=dict(size=11,color=TXS)),
                            xaxis_title="Cₑ (mg/L)",yaxis_title="Cₑ/qₑ",showlegend=True)
                        st.plotly_chart(fig_ll,use_container_width=True)
                    with lt2:
                        # Freundlich linear: ln(qe) vs ln(Ce)
                        mask=Ce>0; lnCe=np.log(Ce[mask]); lnqe=np.log(qe[mask])
                        sl_f,ic_f=np.polyfit(lnCe,lnqe,1)
                        r2_f_lin=r2s(lnqe,sl_f*lnCe+ic_f)
                        fig_fl=go.Figure()
                        fig_fl.add_trace(go.Scatter(x=lnCe,y=lnqe,mode="markers",
                            marker=dict(color=PALETTE[1],size=mk_sz)))
                        xln2=np.linspace(lnCe.min()*.9,lnCe.max()*1.05,100)
                        fig_fl.add_trace(go.Scatter(x=xln2,y=sl_f*xln2+ic_f,mode="lines",
                            line=dict(color=PALETTE[1],width=2),name=f"R²={r2_f_lin:.4f}"))
                        fig_fl.update_layout(**PTBASE,height=220,
                            title=dict(text=t("freundlich_linear"),font=dict(size=11,color=TXS)),
                            xaxis_title="ln Cₑ",yaxis_title="ln qₑ",showlegend=True)
                        st.plotly_chart(fig_fl,use_container_width=True)

                    # D-R & Temkin linear transforms (new row)
                    lt3,lt4=st.columns(2,gap="small")
                    with lt3:
                        # D-R linear: ln(qe) vs ε²  — uses same T & MW as nonlinear fit
                        try:
                            R_kJ=8.314e-3
                            Ce_mol_lt=np.maximum(Ce/(float(mw_dr)*1000.0),1e-15)
                            eps_lt=R_kJ*T_K_iso*np.log(1.0+1.0/Ce_mol_lt)
                            eps2_lt=eps_lt**2
                            lnqe_lt=np.log(np.maximum(qe,1e-15))
                            sl_dr,ic_dr=np.polyfit(eps2_lt,lnqe_lt,1)
                            r2_dr_lin=r2s(lnqe_lt,sl_dr*eps2_lt+ic_dr)
                            beta_lin=-sl_dr if sl_dr<0 else abs(sl_dr)
                            E_lin=1.0/math.sqrt(2.0*beta_lin) if beta_lin>0 else 0.0
                            qm_lin=math.exp(ic_dr)
                            fig_dr_lin=go.Figure()
                            fig_dr_lin.add_trace(go.Scatter(x=eps2_lt,y=lnqe_lt,mode="markers",
                                marker=dict(color=PALETTE[3],size=mk_sz)))
                            xdr=np.linspace(eps2_lt.min(),eps2_lt.max(),100)
                            fig_dr_lin.add_trace(go.Scatter(x=xdr,y=sl_dr*xdr+ic_dr,mode="lines",
                                line=dict(color=PALETTE[3],width=2),
                                name=f"R²={r2_dr_lin:.4f} | E={E_lin:.2f} kJ/mol"))
                            fig_dr_lin.update_layout(**PTBASE,height=220,
                                title=dict(text="D-R Linear: ln(qₑ) vs ε²",font=dict(size=11,color=TXS)),
                                xaxis_title="ε² (kJ²/mol²)",yaxis_title="ln qₑ",showlegend=True)
                            st.plotly_chart(fig_dr_lin,use_container_width=True)
                        except Exception as _e_dr:
                            st.info(f"D-R linear plot: {_e_dr}")
                    with lt4:
                        # Temkin linear: qe vs ln(Ce)
                        try:
                            lnCe_tm=np.log(np.maximum(Ce,1e-15))
                            sl_tm,ic_tm=np.polyfit(lnCe_tm,qe,1)
                            r2_tm_lin=r2s(qe,sl_tm*lnCe_tm+ic_tm)
                            fig_tm_lin=go.Figure()
                            fig_tm_lin.add_trace(go.Scatter(x=lnCe_tm,y=qe,mode="markers",
                                marker=dict(color=PALETTE[2],size=mk_sz)))
                            xtm=np.linspace(lnCe_tm.min(),lnCe_tm.max(),100)
                            fig_tm_lin.add_trace(go.Scatter(x=xtm,y=sl_tm*xtm+ic_tm,mode="lines",
                                line=dict(color=PALETTE[2],width=2),name=f"R²={r2_tm_lin:.4f}"))
                            fig_tm_lin.update_layout(**PTBASE,height=220,
                                title=dict(text="Temkin Linear: qₑ vs ln(Cₑ)",font=dict(size=11,color=TXS)),
                                xaxis_title="ln Cₑ",yaxis_title="qₑ (mg/g)",showlegend=True)
                            st.plotly_chart(fig_tm_lin,use_container_width=True)
                        except Exception as _e_tm:
                            st.info(f"Temkin linear plot: {_e_tm}")

                    if show_res:
                        st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('residual_plot')}</b>",unsafe_allow_html=True)
                        fig_r=go.Figure()
                        for i,(nm,res) in enumerate(iso_res.items()):
                            try:
                                yp=funcs[nm](Ce,*res["popt"])
                                fig_r.add_trace(go.Scatter(x=Ce,y=qe-yp,mode="markers+lines",
                                    name=nm,marker=dict(color=PALETTE[i],size=8),
                                    line=dict(color=PALETTE[i],width=1,dash="dot")))
                            except: pass
                        fig_r.add_hline(y=0,line_dash="dash",line_color="rgba(148,163,184,0.3)")
                        fig_r.update_layout(**PTBASE,height=200,
                            xaxis_title="Cₑ",yaxis_title="Residuals")
                        st.plotly_chart(fig_r,use_container_width=True)

                    st.markdown(f"<b style='color:{TXS};font-size:.86rem;font-family:{ff};'>{t('computed_data')}</b>",unsafe_allow_html=True)
                    dd=df_i[["C0","Absorbance","Ce","qe","Rem%"]].copy().round(4)
                    dd.columns=["C₀","Abs","Cₑ (mg/L)","qₑ (mg/g)","Rem%"]
                    st.dataframe(dd,use_container_width=True,height=190)

                with st_col:
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('model_ranking')}</b>",unsafe_allow_html=True)
                    icons=["🥇","🥈","🥉","🔹","🔸","▪️"]
                    srt=sorted(iso_res.items(),key=lambda x:-x[1]["r2"])
                    for rk,(nm,res) in enumerate(srt):
                        ib=(rk==0); ic=icons[min(rk,5)]
                        _cls="ada-card best" if ib else "ada-card"
                        _clr="#34d399" if ib else AC
                        _bdg=f"<span class='ada-badge'>{t('best_fit')}</span>" if ib else ""
                        _r2=f"{res['r2']:.4f}"; _rm=f"{res['rmse']:.4f}"; _ch=f"{res['chi2']:.4f}"
                        # Model-specific favorability indicator
                        _fav=""
                        if nm=="Langmuir":
                            RL_v=res["params"].get("RL",0.5)
                            if RL_v<=0:   _fl,_fc="Irreversible","#a78bfa"
                            elif RL_v<1:  _fl,_fc=f"Favorable \u2714 (RL={RL_v:.3f})","#34d399"
                            elif RL_v==1: _fl,_fc=f"Linear (RL={RL_v:.3f})","#fbbf24"
                            else:         _fl,_fc=f"Unfavorable \u26a0 (RL={RL_v:.3f})","#f87171"
                            _fav=f'<div style="margin-top:.3rem;font-size:.62rem;color:{_fc};font-family:monospace;">\u25cf {_fl}</div>'
                        elif nm=="Freundlich":
                            inv_n=res["params"].get("1/n",1.0)
                            if inv_n<1:   _fl,_fc=f"Favorable \u2714 (1/n={inv_n:.3f})","#34d399"
                            elif inv_n>1: _fl,_fc=f"Unfavorable \u26a0 (1/n={inv_n:.3f})","#f87171"
                            else:         _fl,_fc=f"Linear (1/n={inv_n:.3f})","#fbbf24"
                            _fav=f'<div style="margin-top:.3rem;font-size:.62rem;color:{_fc};font-family:monospace;">\u25cf {_fl}</div>'
                        elif nm=="D-R":
                            E_v=res["params"].get("E (kJ/mol)",0); ads_t=res["params"].get("Type","")
                            _fc="#60a5fa" if "Physi" in ads_t else ("#fbbf24" if "Ion" in ads_t else "#f87171")
                            _fav=f'<div style="margin-top:.3rem;font-size:.62rem;color:{_fc};font-family:monospace;">\u25cf {ads_t} (E={E_v:.1f} kJ/mol)</div>'
                        elif nm=="Sips":
                            ns_v=res["params"].get("ns",1.0)
                            if abs(ns_v-1)<0.15: _fl,_fc=f"\u2248Langmuir (ns={ns_v:.3f})","#fbbf24"
                            else:                _fl,_fc=f"Heterogeneous (ns={ns_v:.3f})","#34d399"
                            _fav=f'<div style="margin-top:.3rem;font-size:.62rem;color:{_fc};font-family:monospace;">\u25cf {_fl}</div>'
                        elif nm=="Temkin":
                            bT_v=res["params"].get("bT (J/mol)",0)
                            _fl="Physisorption" if bT_v<40000 else "Chemisorption"
                            _fc="#60a5fa" if bT_v<40000 else "#f87171"
                            _fav=f'<div style="margin-top:.3rem;font-size:.62rem;color:{_fc};font-family:monospace;">\u25cf {_fl} (bT={bT_v:.0f} J/mol)</div>'
                        st.markdown(
                            f'<div class="{_cls}">'
                            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:.35rem;">'
                            f'<span style="font-weight:700;color:{TX};font-size:.87rem;font-family:{ff};">{ic} {nm}</span>{_bdg}'
                            f'</div>'
                            f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:.25rem;">'
                            f'<div><div style="color:{TXM};font-size:.58rem;">R\u00b2</div>'
                            f'<div style="color:{_clr};font-family:monospace;font-size:.81rem;font-weight:700;">{_r2}</div></div>'
                            f'<div><div style="color:{TXM};font-size:.58rem;">RMSE</div>'
                            f'<div style="color:{TXS};font-family:monospace;font-size:.81rem;">{_rm}</div></div>'
                            f'<div><div style="color:{TXM};font-size:.58rem;">\u03c7\u00b2</div>'
                            f'<div style="color:{TXS};font-family:monospace;font-size:.81rem;">{_ch}</div></div>'
                            f'</div>{_fav}</div>',
                            unsafe_allow_html=True)

                    # Data quality
                    st.markdown(f"<br><b style='color:{TXS};font-family:{ff};'>{t('data_quality')}</b>",unsafe_allow_html=True)
                    st.markdown(f"""
                    <div class="ada-card">
                        <div class="param-row"><span class="param-key">{t("dq_cv")}</span><span class="param-val">{cv_val:.1f}%</span></div>
                        <div class="param-row"><span class="param-key">{t("dq_range")}</span><span class="param-val">{qe.min():.2f}–{qe.max():.2f}</span></div>
                        <div class="param-row"><span class="param-key">{t("dq_pts")}</span><span class="param-val">{len(df_i)}</span></div>
                        <div class="param-row"><span class="param-key">{t("dq_mono")}</span><span class="param-val">{"✅" if mono else "⚠️"}</span></div>
                    </div>""",unsafe_allow_html=True)

                    # Parameters with 95% CI
                    st.markdown(f"<br><b style='color:{TXS};font-family:{ff};'>{t('parameters')}</b>",unsafe_allow_html=True)
                    for nm,res in srt:
                        with st.expander(nm,expanded=(nm==best_iso)):
                            popt_list=list(res["popt"])
                            pcov_m=res.get("pcov")
                            for i,(pk,pv) in enumerate(res["params"].items()):
                                if isinstance(pv,float):
                                    ci_str=""
                                    if pcov_m is not None and i<len(popt_list):
                                        try:
                                            ci=1.96*float(np.sqrt(abs(pcov_m[i,i])))
                                            if np.isfinite(ci): ci_str=f" <span style='color:{TXM};font-size:.7rem;'>±{ci:.4f}</span>"
                                        except: pass
                                    st.markdown(f'<div class="param-row"><span class="param-key">{pk}</span><span class="param-val">{pv:.4f}{ci_str}</span></div>',unsafe_allow_html=True)
                                else:
                                    st.markdown(f'<div class="param-row"><span class="param-key">{pk}</span><span class="param-val">{pv}</span></div>',unsafe_allow_html=True)

                    # Auto interpretation
                    st.markdown(f"<br><b style='color:{TXS};font-family:{ff};'>{t('auto_interp')}</b>",unsafe_allow_html=True)
                    interp_lines=[]
                    interp_lines.append(f"• {t('interp_best')} <b>{best_iso}</b> {t('interp_r2')} {iso_res[best_iso]['r2']:.4f}")
                    if "Langmuir" in iso_res:
                        RL_v=iso_res["Langmuir"]["params"].get("RL",0.5)
                        if isinstance(RL_v,float):
                            interp_lines.append(f"• {t('interp_lang_fav') if 0<RL_v<1 else t('interp_lang_unfav')} (RL={RL_v:.3f})")
                    if "Freundlich" in iso_res:
                        inv_n=iso_res["Freundlich"]["params"].get("1/n",1.)
                        if isinstance(inv_n,float):
                            interp_lines.append(f"• {t('interp_frnd_good') if inv_n<1 else t('interp_frnd_poor')} (1/n={inv_n:.3f})")
                    if "D-R" in iso_res:
                        Ev=iso_res["D-R"]["params"].get("E (kJ/mol)",0)
                        if isinstance(Ev,float):
                            interp_lines.append(f"• D-R: E = {Ev:.2f} kJ/mol → {'Physical' if Ev<8 else 'Chemical'} adsorption")
                    if "Sips" in iso_res:
                        ns_v=iso_res["Sips"]["params"].get("ns",1.)
                        if isinstance(ns_v,float):
                            interp_lines.append(f"• Sips ns = {ns_v:.3f} ({'≈Langmuir' if abs(ns_v-1)<.15 else 'Heterogeneous'})")
                    st.markdown(f"""
                    <div class="interp-box">
                        {"".join([f'<div class="interp-line">{l}</div>' for l in interp_lines])}
                    </div>""",unsafe_allow_html=True)

                st.session_state.update({"iso_results":iso_res,"iso_df":dd,"iso_best":best_iso})

                # ── Prediction Tool ────────────────────────────────────────
                st.markdown(f"<hr style='border:none;border-top:1px solid {BR};margin:1rem 0;'>",unsafe_allow_html=True)
                with st.expander(f"🔮 qₑ Prediction Calculator",expanded=False):
                    st.markdown(f'<div class="ada-eq">📐 Enter a Cₑ value → get predicted qₑ from every fitted model</div>',unsafe_allow_html=True)
                    st.markdown("<br>",unsafe_allow_html=True)
                    pred_ce=st.number_input("Enter Cₑ (mg/L)",value=float(round(Ce.mean(),3)),min_value=1e-6,format="%.4f",key="pred_ce")
                    if pred_ce>0:
                        st.markdown(f"<b style='color:{TXS};font-family:{ff};font-size:.87rem;'>Predicted qₑ (mg/g) at Cₑ = {pred_ce:.4f} mg/L</b>",unsafe_allow_html=True)
                        for nm,res in sorted(iso_res.items(),key=lambda x:-x[1]["r2"]):
                            try:
                                q_pred=float(funcs[nm](np.array([pred_ce]),*res["popt"])[0])
                                ib=(nm==best_iso)
                                if np.isfinite(q_pred):
                                    st.markdown(f'<div class="param-row"><span class="param-key" style="font-weight:{"700" if ib else "400"};color:{TX if ib else TXS};">{"★ " if ib else ""}{nm}</span><span class="param-val" style="color:{"#34d399" if ib else AC};">{q_pred:.4f} mg/g</span></div>',unsafe_allow_html=True)
                            except: pass

        except Exception as e: st.error(f"Error: {e}")
    else: st.info(t("iso_min_pts"))

# ─────────────────────────────────────────────────────────────────────────────
#  KINETICS
# ─────────────────────────────────────────────────────────────────────────────
with tab_kin:
    st.markdown(f"""
    <div class="ada-section"><div class="ada-icon">⏱️</div>
    <div><div class="ada-title">{t("kin_header")}</div>
    <div class="ada-desc">{t("kin_desc")}</div></div></div>""",unsafe_allow_html=True)

    kc1,kc2,kc3=st.columns(3)
    with kc1: c0k=st.number_input(t("c0_kin"),value=50.0)
    with kc2: vk=st.number_input(t("vol"),value=.050,format="%.3f",key="vk")
    with kc3: mk=st.number_input(t("mass"),value=.100,format="%.3f",key="mk")
    kin_sel=st.multiselect(t("models_fit"),["PFO","PSO","Elovich","Weber-Morris"],
        default=["PFO","PSO","Elovich","Weber-Morris"],key="ksel")

    st.markdown(f"<hr style='border:none;border-top:1px solid {BR};margin:.5rem 0;'>",unsafe_allow_html=True)
    mkd=st.radio(t("data_input"),[t("upload"),t("manual")],horizontal=True,key="mkin")

    if mkd==t("upload"):
        uf2=st.file_uploader(t("upload_hint_kin"),type=["xlsx","csv"],key="kuf")
        if uf2:
            try:
                dr2=pd.read_excel(uf2) if uf2.name.endswith(".xlsx") else pd.read_csv(uf2)
                if "Time" in dr2.columns and "Absorbance" in dr2.columns:
                    st.session_state.kin_data=dr2[["Time","Absorbance"]].copy()
                else: st.error(t("error_cols_kin"))
            except Exception as e: st.error(str(e))
    else:
        st.session_state.kin_data=st.data_editor(st.session_state.kin_data,num_rows="dynamic",
            use_container_width=True,key="ked",
            column_config={"Time":st.column_config.NumberColumn("Time (min)",format="%.1f"),
                           "Absorbance":st.column_config.NumberColumn("Absorbance",format="%.4f")})

    df_k=st.session_state.kin_data.copy()
    if df_k is not None and len(df_k)>=4:
        try:
            df_k["Ct"]=get_c(df_k["Absorbance"].values,cal_key,**cal_kw)
            df_k["qt"]=((c0k-df_k["Ct"])*vk)/mk
            df_k=df_k[df_k["qt"]>=0].reset_index(drop=True)
            Td,qt=df_k["Time"].values,df_k["qt"].values

            kp1,kp2,kp3,kp4=st.columns(4)
            qt_last = float(qt[-1]) if len(qt)>0 else 0.0
            td_last = float(Td[-1]) if len(Td)>0 else 0.0
            for col,(val,lbl,sub) in zip([kp1,kp2,kp3,kp4],[
                (f"{qt_last:.3f}",t("qt_max"),t("duration")),
                (f"{td_last:.0f} min",t("max_time"),t("duration")),
                (f"{qt.max():.3f}",t("max_qt"),t("peak_kin")),
                (f"{len(df_k)}",t("data_pts"),t("valid_meas"))]):
                with col:
                    st.markdown(f"""<div class="ada-metric"><div class="ada-metric-val">{val}</div>
                        <div class="ada-metric-lbl">{lbl}</div><div class="ada-metric-sub">{sub}</div></div>""",unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)

            kfuncs={"PFO":pfo,"PSO":pso,"Elovich":elovich,"Weber-Morris":wm}
            kp0s={"PFO":[qt.max(),.05],"PSO":[qt.max(),.01],
                  "Elovich":[1.,.5],"Weber-Morris":[qt.max()/np.sqrt(Td.max()+1e-6),.1]}
            kbds={"PFO":(0,np.inf),"PSO":(0,np.inf),
                  "Elovich":([1e-6,1e-6],[np.inf,np.inf]),"Weber-Morris":(-np.inf,np.inf)}
            kpms={"PFO": lambda p:{"qₑ (mg/g)":p[0],"k₁ (min⁻¹)":p[1],"t½ (min)":math.log(2)/p[1] if p[1]>0 else float("inf")},
                  "PSO": lambda p:{"qₑ (mg/g)":p[0],"k₂ (g/mg·min)":p[1],"h₀ (mg/g·min)":p[1]*p[0]**2},
                  "Elovich":lambda p:{"α (mg/g·min)":p[0],"β (g/mg)":p[1]},
                  "Weber-Morris":lambda p:{"k_id":p[0],"C (mg/g)":p[1]}}

            kin_res={}
            for nm in kin_sel:
                po,pcov_k,rv,rm,rc_k=do_fit(kfuncs[nm],Td,qt,kp0s[nm],kbds[nm])
                if po is not None:
                    kin_res[nm]={"popt":po,"pcov":pcov_k,"r2":rv,"rmse":rm,"chi2":rc_k,"params":kpms[nm](po)}

            if kin_res:
                best_kin=max(kin_res,key=lambda k:kin_res[k]["r2"])
                kch,kst=st.columns([3,2],gap="large")

                with kch:
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('kin_curves')}</b>",unsafe_allow_html=True)
                    fig_k=go.Figure()
                    fig_k.add_trace(go.Scatter(x=Td,y=qt,mode="markers",name="Exp.",
                        marker=dict(color="#f87171",size=mk_sz,symbol="diamond",
                                    line=dict(color="rgba(248,113,113,0.3)",width=1))))
                    tL=np.linspace(.01,Td.max()*1.1,300)
                    for i,(nm,res) in enumerate(kin_res.items()):
                        yL=kfuncs[nm](tL,*res["popt"]); ib=(nm==best_kin)
                        fig_k.add_trace(go.Scatter(x=tL,y=yL,mode="lines",
                            name=f"{nm} R²={res['r2']:.4f}"+(" ★" if ib else ""),
                            line=dict(color=PALETTE[i],width=3 if ib else 1.8,dash="solid" if ib else "dot")))
                    fig_k.update_layout(**PTBASE,height=340,xaxis_title="Time (min)",yaxis_title="qt (mg/g)")
                    st.plotly_chart(fig_k,use_container_width=True)

                    # R² bar
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('r2_chart')}</b>",unsafe_allow_html=True)
                    knms=[nm for nm in kin_res]; kr2=[kin_res[nm]["r2"] for nm in knms]
                    fig_kb=go.Figure(go.Bar(x=knms,y=kr2,marker_color=PALETTE[:len(knms)],
                        text=[f"{v:.4f}" for v in kr2],textposition="outside"))
                    fig_kb.update_layout(**PTBASE,height=200,showlegend=False)
                    fig_kb.update_yaxes(range=[max(0,min(kr2)-.05),1.02],gridcolor=PG)
                    st.plotly_chart(fig_kb,use_container_width=True)

                    if "Weber-Morris" in kin_res:
                        st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('wm_plot')}</b>",unsafe_allow_html=True)
                        sT=np.sqrt(Td); fig_wm=go.Figure()
                        fig_wm.add_trace(go.Scatter(x=sT,y=qt,mode="markers",
                            marker=dict(color="#fb923c",size=mk_sz)))
                        pwm=kin_res["Weber-Morris"]["popt"]
                        xwm=np.linspace(0,sT.max()*1.1,200)
                        fig_wm.add_trace(go.Scatter(x=xwm,y=wm(xwm,*pwm),mode="lines",
                            name=f"R²={kin_res['Weber-Morris']['r2']:.4f}",
                            line=dict(color="#fb923c",width=2)))
                        fig_wm.add_hline(y=pwm[1],line_dash="dash",
                            line_color="rgba(251,146,60,.35)",
                            annotation_text=f"C={pwm[1]:.3f}")
                        fig_wm.update_layout(**PTBASE,height=230,
                            xaxis_title="√t (min⁰·⁵)",yaxis_title="qt (mg/g)")
                        st.plotly_chart(fig_wm,use_container_width=True)

                    # ── Kinetics Linear Transform Plots ───────────────────────
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>📐 Kinetics Linear Transforms</b>",unsafe_allow_html=True)
                    klt1,klt2=st.columns(2,gap="small")
                    with klt1:
                        # PFO linear: ln(qe-qt) vs t → slope=-k1, intercept=ln(qe)
                        try:
                            qe_pfo=kin_res["PFO"]["popt"][0] if "PFO" in kin_res else qt.max()
                            diff=qe_pfo-qt
                            mask_pfo=diff>0
                            if mask_pfo.sum()>=3:
                                lnD=np.log(diff[mask_pfo]); td_pfo=Td[mask_pfo]
                                sl_p1,ic_p1=np.polyfit(td_pfo,lnD,1)
                                r2_p1=r2s(lnD,sl_p1*td_pfo+ic_p1)
                                k1_lin=-sl_p1; qe_lin=math.exp(ic_p1)
                                fig_pfo_lin=go.Figure()
                                fig_pfo_lin.add_trace(go.Scatter(x=td_pfo,y=lnD,mode="markers",
                                    marker=dict(color=PALETTE[0],size=mk_sz)))
                                xt_p=np.linspace(td_pfo.min(),td_pfo.max(),100)
                                fig_pfo_lin.add_trace(go.Scatter(x=xt_p,y=sl_p1*xt_p+ic_p1,mode="lines",
                                    line=dict(color=PALETTE[0],width=2),
                                    name=f"R²={r2_p1:.4f} | k₁={k1_lin:.4f}"))
                                fig_pfo_lin.update_layout(**PTBASE,height=230,
                                    title=dict(text="PFO Linear: ln(qₑ−qt) vs t",font=dict(size=11,color=TXS)),
                                    xaxis_title="t (min)",yaxis_title="ln(qₑ−qt)",showlegend=True)
                                st.plotly_chart(fig_pfo_lin,use_container_width=True)
                            else:
                                st.info("PFO linear: insufficient valid points (qe−qt must be >0)")
                        except Exception as _ep: st.info(f"PFO linear: {_ep}")
                    with klt2:
                        # PSO linear: t/qt vs t → slope=1/qe, intercept=1/(k2*qe²)
                        try:
                            mask_pso=qt>0
                            if mask_pso.sum()>=3:
                                tqt=Td[mask_pso]/qt[mask_pso]; td_pso=Td[mask_pso]
                                sl_p2,ic_p2=np.polyfit(td_pso,tqt,1)
                                r2_p2=r2s(tqt,sl_p2*td_pso+ic_p2)
                                qe_p2=1.0/sl_p2 if sl_p2>0 else 0.0
                                k2_lin=sl_p2**2/ic_p2 if ic_p2>0 else 0.0
                                fig_pso_lin=go.Figure()
                                fig_pso_lin.add_trace(go.Scatter(x=td_pso,y=tqt,mode="markers",
                                    marker=dict(color=PALETTE[1],size=mk_sz)))
                                xt_p2=np.linspace(td_pso.min(),td_pso.max(),100)
                                fig_pso_lin.add_trace(go.Scatter(x=xt_p2,y=sl_p2*xt_p2+ic_p2,mode="lines",
                                    line=dict(color=PALETTE[1],width=2),
                                    name=f"R²={r2_p2:.4f} | qₑ={qe_p2:.3f}"))
                                fig_pso_lin.update_layout(**PTBASE,height=230,
                                    title=dict(text="PSO Linear: t/qt vs t",font=dict(size=11,color=TXS)),
                                    xaxis_title="t (min)",yaxis_title="t/qt (min·g/mg)",showlegend=True)
                                st.plotly_chart(fig_pso_lin,use_container_width=True)
                        except Exception as _ep2: st.info(f"PSO linear: {_ep2}")

                with kst:
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('model_ranking')}</b>",unsafe_allow_html=True)
                    icons=["🥇","🥈","🥉","🔹"]
                    for rk,(nm,res) in enumerate(sorted(kin_res.items(),key=lambda x:-x[1]["r2"])):
                        ib=(rk==0); ic=icons[min(rk,3)]
                        _cls="ada-card best" if ib else "ada-card"
                        _clr="#34d399" if ib else AC
                        _bdg=f"<span class='ada-badge'>{t('best_fit')}</span>" if ib else ""
                        _r2=f"{res['r2']:.4f}"; _rm=f"{res['rmse']:.4f}"; _ch=f"{res['chi2']:.4f}"
                        st.markdown(
                            f'<div class="{_cls}">'
                            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:.35rem;">'
                            f'<span style="font-weight:700;color:{TX};font-size:.87rem;font-family:{ff};">{ic} {nm}</span>{_bdg}'
                            f'</div>'
                            f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:.25rem;">'
                            f'<div><div style="color:{TXM};font-size:.58rem;">R\u00b2</div>'
                            f'<div style="color:{_clr};font-family:monospace;font-size:.81rem;font-weight:700;">{_r2}</div></div>'
                            f'<div><div style="color:{TXM};font-size:.58rem;">RMSE</div>'
                            f'<div style="color:{TXS};font-family:monospace;font-size:.81rem;">{_rm}</div></div>'
                            f'<div><div style="color:{TXM};font-size:.58rem;">\u03c7\u00b2</div>'
                            f'<div style="color:{TXS};font-family:monospace;font-size:.81rem;">{_ch}</div></div>'
                            f'</div></div>',
                            unsafe_allow_html=True)

                    st.markdown(f"<br><b style='color:{TXS};font-family:{ff};'>{t('parameters')}</b>",unsafe_allow_html=True)
                    for nm,res in sorted(kin_res.items(),key=lambda x:-x[1]["r2"]):
                        with st.expander(nm,expanded=(nm==best_kin)):
                            popt_list_k=list(res["popt"])
                            pcov_mk=res.get("pcov")
                            for i,(pk,pv) in enumerate(res["params"].items()):
                                ci_str_k=""
                                if pcov_mk is not None and i<len(popt_list_k):
                                    try:
                                        ci_k=1.96*float(np.sqrt(abs(pcov_mk[i,i])))
                                        if np.isfinite(ci_k): ci_str_k=f" <span style='color:{TXM};font-size:.7rem;'>±{ci_k:.4f}</span>"
                                    except: pass
                                st.markdown(f'<div class="param-row"><span class="param-key">{pk}</span><span class="param-val">{pv:.4f}{ci_str_k}</span></div>',unsafe_allow_html=True)

                    # Auto interp
                    st.markdown(f"<br><b style='color:{TXS};font-family:{ff};'>{t('auto_interp')}</b>",unsafe_allow_html=True)
                    klines=[f"• {t('interp_best')} <b>{best_kin}</b> {t('interp_r2')} {kin_res[best_kin]['r2']:.4f}"]
                    if "PSO" in kin_res and kin_res["PSO"]["r2"]>kin_res.get("PFO",{"r2":0})["r2"]:
                        klines.append("• PSO > PFO → chemical interaction dominant")
                    else:
                        klines.append("• PFO ≥ PSO → physical diffusion dominant")
                    if "Weber-Morris" in kin_res:
                        C_wm=kin_res["Weber-Morris"]["params"].get("C (mg/g)",0)
                        klines.append(f"• WM C = {C_wm:.3f} → {'film diffusion' if C_wm>0.1 else 'intraparticle diffusion'}")
                    st.markdown(f"""
                    <div class="interp-box">
                        {"".join([f'<div class="interp-line">{l}</div>' for l in klines])}
                    </div>""",unsafe_allow_html=True)

                st.session_state.update({"kin_results":kin_res,"kin_df":df_k[["Time","Absorbance","Ct","qt"]].round(4),"kin_best":best_kin})

            # ── qt Prediction Calculator ────────────────────────────────────
            if st.session_state.get("kin_results"):
                st.markdown(f"<hr style='border:none;border-top:1px solid {BR};margin:1rem 0;'>",unsafe_allow_html=True)
                with st.expander("🧮 qt Prediction Calculator — predict adsorption capacity at any time",expanded=False):
                    st.markdown(f'<div class="ada-eq">📐 Enter a contact time → get predicted qt (mg/g) from every fitted kinetic model</div>',unsafe_allow_html=True)
                    st.markdown("<br>",unsafe_allow_html=True)
                    t_pred_val=st.number_input("Contact time t (min)",min_value=0.0,value=30.0,format="%.2f",key="kin_t_pred")
                    kr=st.session_state["kin_results"]
                    pred_rows=[]
                    for mnm,mdata in kr.items():
                        po_k=mdata["popt"]
                        try:
                            if mnm=="PFO":
                                qt_pred=pfo(t_pred_val,*po_k)
                            elif mnm=="PSO":
                                qt_pred=pso(t_pred_val,*po_k)
                            elif mnm=="Elovich":
                                qt_pred=elovich(t_pred_val,*po_k)
                            elif mnm=="Weber-Morris":
                                qt_pred=wm(t_pred_val,*po_k)
                            else:
                                continue
                            pred_rows.append({"Model":mnm,"qt predicted (mg/g)":round(float(qt_pred),4),"R²":round(mdata["r2"],4)})
                        except Exception:
                            pass
                    if pred_rows:
                        df_kpred=pd.DataFrame(pred_rows).sort_values("R²",ascending=False).reset_index(drop=True)
                        best_row=df_kpred.iloc[0]
                        qpc1,qpc2=st.columns([2,1])
                        with qpc1:
                            st.dataframe(df_kpred,use_container_width=True,hide_index=True)
                        with qpc2:
                            st.markdown(f"""<div class="ada-metric" style="border-left:4px solid #22c55e;">
                                <div class="ada-metric-val" style="color:#16a34a;">{best_row['qt predicted (mg/g)']:.4f}</div>
                                <div class="ada-metric-lbl">mg/g @ t={t_pred_val:.1f} min</div>
                                <div class="ada-metric-sub">Best model: {best_row['Model']}</div></div>""",unsafe_allow_html=True)

            # ── Arrhenius Section ──────────────────────────────────────────
            st.markdown(f"<hr style='border:none;border-top:1px solid {BR};margin:1rem 0;'>",unsafe_allow_html=True)
            with st.expander(f"🔥 {t('arrhenius_title')} — {t('arrhenius_desc')}",expanded=False):
                st.markdown(f'<div class="ada-eq">📐 {t("arrhenius_exp")}</div>',unsafe_allow_html=True)
                st.markdown("<br>",unsafe_allow_html=True)
                na=st.number_input(t("num_arr_temps"),min_value=2,max_value=6,value=3,key="narr")
                arr_data=[]
                acols=st.columns(int(na))
                dT=[25,35,45,55,65,75]; dk=[0.05,0.12,0.28,0.55,0.95,1.60]
                for i,col in enumerate(acols):
                    with col:
                        aT=st.number_input(t("arr_temp"),value=float(dT[i]),key=f"at{i}")
                        ak=st.number_input(t("arr_k"),value=float(dk[i]),format="%.4f",key=f"ak{i}")
                        arr_data.append({"T_C":aT,"k":ak})
                df_arr=pd.DataFrame(arr_data)
                df_arr["T_K"]=df_arr["T_C"]+273.15
                df_arr["invT"]=1/df_arr["T_K"]
                df_arr["lnk"]=np.log(df_arr["k"])
                R=8.314
                try:
                    sl_a,ic_a=np.polyfit(df_arr["invT"],df_arr["lnk"],1)
                    Ea=-sl_a*R/1000
                    r2_arr=r2s(df_arr["lnk"].values,(sl_a*df_arr["invT"]+ic_a).values)

                    ac1,ac2,ac3=st.columns(3)
                    for col,(val,lbl) in zip([ac1,ac2,ac3],[
                        (f"{Ea:.2f} kJ/mol",t("ea_result")),
                        (f"{r2_arr:.4f}",t("arr_r2")),
                        ("Physical" if Ea<40 else "Chemical","Type")]):
                        with col:
                            st.markdown(f"""<div class="ada-metric"><div class="ada-metric-val" style="font-size:1.3rem;">{val}</div>
                                <div class="ada-metric-lbl">{lbl}</div></div>""",unsafe_allow_html=True)

                    st.markdown("<br>",unsafe_allow_html=True)
                    al,ar_=st.columns(2,gap="large")
                    with al:
                        invL=np.linspace(df_arr["invT"].min()*.99,df_arr["invT"].max()*1.01,200)
                        fig_arr=go.Figure()
                        fig_arr.add_trace(go.Scatter(x=invL,y=sl_a*invL+ic_a,mode="lines",
                            name=f"Arrhenius Fit (R²={r2_arr:.4f})",line=dict(color=AC,width=2.5)))
                        fig_arr.add_trace(go.Scatter(x=df_arr["invT"],y=df_arr["lnk"],mode="markers",
                            name="Exp.",marker=dict(color="#f87171",size=mk_sz+2)))
                        fig_arr.update_layout(**PTBASE,height=260,
                            xaxis_title="1/T (K⁻¹)",yaxis_title="ln k")
                        st.plotly_chart(fig_arr,use_container_width=True)
                    with ar_:
                        st.markdown(f"""
                        <div class="interp-box" style="margin-top:.5rem;">
                            <div class="interp-title">🔥 Activation Energy Analysis</div>
                            <div class="interp-line">• Eₐ = {Ea:.2f} kJ/mol</div>
                            <div class="interp-line">• {"<b>"+t("ea_interp_phys")+"</b>" if Ea<40 else "<b>"+t("ea_interp_chem")+"</b>"}</div>
                            <div class="interp-line">• Arrhenius R² = {r2_arr:.4f}</div>
                            <div class="interp-line">• Pre-exponential A = {math.exp(ic_a):.4f}</div>
                        </div>""",unsafe_allow_html=True)
                except Exception as e: st.error(str(e))

        except Exception as e: st.error(f"Error: {e}")
    else: st.info(t("kin_min_pts"))

# ─────────────────────────────────────────────────────────────────────────────
#  THERMODYNAMICS
# ─────────────────────────────────────────────────────────────────────────────
with tab_thermo:
    st.markdown(f"""
    <div class="ada-section"><div class="ada-icon">🌡️</div>
    <div><div class="ada-title">{t("thermo_header")}</div>
    <div class="ada-desc">{t("thermo_desc")}</div></div></div>""",unsafe_allow_html=True)

    st.markdown(f'<div class="ada-eq">📐 Van\'t Hoff: ln(Kc) = −ΔH°/RT + ΔS°/R &nbsp;|&nbsp; ΔG° = −RT·ln(Kc)</div>',unsafe_allow_html=True)
    st.markdown("<br>",unsafe_allow_html=True)

    nt=st.number_input(t("num_temps"),min_value=2,max_value=8,value=3)
    dTT=[25,35,45,55,65,75,85,95]; dKK=[2.,3.2,4.8,6.5,8.2,10.,12.5,15.]
    thd=[]
    tcols=st.columns(int(nt))
    for i,col in enumerate(tcols):
        with col:
            tc=st.number_input(t("temp_pt"),value=float(dTT[i]),key=f"tc{i}")
            kc=st.number_input(t("kc_pt"),value=float(dKK[i]),format="%.4f",key=f"kc{i}")
            thd.append({"T_C":tc,"Kc":kc})

    df_th=pd.DataFrame(thd)
    df_th["T_K"]=df_th["T_C"]+273.15
    df_th["invT"]=1/df_th["T_K"]
    df_th["lnKc"]=np.log(df_th["Kc"])
    R=8.314
    try:
        slt,ict=np.polyfit(df_th["invT"],df_th["lnKc"],1)
        dH=-slt*R/1000; dS=ict*R
        df_th["dG"]=(-R*df_th["T_K"]*df_th["lnKc"])/1000
        r2_th=r2s(df_th["lnKc"].values,(slt*df_th["invT"]+ict).values)

        tm1,tm2,tm3,tm4=st.columns(4)
        for col,(val,lbl,sub) in zip([tm1,tm2,tm3,tm4],[
            (f"{dH:.3f} kJ/mol",t("enthalpy"),t("exo_desc")),
            (f"{dS:.3f} J/mol·K",t("entropy"),t("entropy_desc")),
            (f"{df_th['dG'].iloc[0]:.3f} kJ/mol",t("dg_at_t1"),"kJ/mol"),
            (f"{r2_th:.4f}",t("vth_r2"),t("linear_fit"))]):
            with col:
                st.markdown(f"""<div class="ada-metric" style="margin-top:.5rem;">
                    <div class="ada-metric-val" style="font-size:1.2rem;">{val}</div>
                    <div class="ada-metric-lbl">{lbl}</div><div class="ada-metric-sub">{sub}</div></div>""",unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)
        th1,th2=st.columns([3,2],gap="large")

        with th1:
            st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('vth_plot')}</b>",unsafe_allow_html=True)
            invL=np.linspace(df_th["invT"].min()*.99,df_th["invT"].max()*1.01,200)
            fig_vth=go.Figure()
            fig_vth.add_trace(go.Scatter(x=invL,y=slt*invL+ict,mode="lines",
                name=f"Van't Hoff (R²={r2_th:.4f})",line=dict(color=AC,width=2.5)))
            fig_vth.add_trace(go.Scatter(x=df_th["invT"],y=df_th["lnKc"],mode="markers",
                name="Exp.",marker=dict(color="#f87171",size=mk_sz+2)))
            fig_vth.update_layout(**PTBASE,height=285,xaxis_title="1/T (K⁻¹)",yaxis_title="ln(Kc)")
            st.plotly_chart(fig_vth,use_container_width=True)

            st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('dg_plot')}</b>",unsafe_allow_html=True)
            bc=[("#34d399" if v<0 else "#f87171") for v in df_th["dG"]]
            fig_dg=go.Figure(go.Bar(
                x=[f"{r['T_C']:.0f}°C" for _,r in df_th.iterrows()],
                y=df_th["dG"],marker=dict(color=bc,line=dict(width=0)),
                text=[f"{v:.2f}" for v in df_th["dG"]],textposition="outside"))
            fig_dg.add_hline(y=0,line_dash="dash",line_color="rgba(148,163,184,.3)")
            fig_dg.update_layout(**PTBASE,height=230,showlegend=False,
                xaxis_title="Temperature",yaxis_title="ΔG° (kJ/mol)")
            st.plotly_chart(fig_dg,use_container_width=True)

            # ΔG° trend line
            st.markdown(f"<b style='color:{TXS};font-family:{ff};'>📉 ΔG° Trend</b>",unsafe_allow_html=True)
            fig_dgt=go.Figure()
            fig_dgt.add_trace(go.Scatter(x=df_th["T_K"],y=df_th["dG"],mode="lines+markers",
                line=dict(color=AC,width=2.5),marker=dict(color=AC,size=mk_sz),name="ΔG°"))
            fig_dgt.add_hline(y=0,line_dash="dash",line_color="rgba(148,163,184,.3)")
            fig_dgt.update_layout(**PTBASE,height=210,
                xaxis_title="T (K)",yaxis_title="ΔG° (kJ/mol)",showlegend=False)
            st.plotly_chart(fig_dgt,use_container_width=True)

        with th2:
            st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('thermo_interp')}</b>",unsafe_allow_html=True)
            if dH<0:
                st.markdown(f"""<div class="ada-card exo">
                    <div style="font-size:.88rem;font-weight:700;color:#fb923c;">{t("exo_title")}</div>
                    <div style="color:{TXS};font-size:.77rem;margin-top:.3rem;line-height:1.6;">
                        ΔH° = {dH:.3f} kJ/mol<br>{t("exo_body")}</div></div>""",unsafe_allow_html=True)
            else:
                st.markdown(f"""<div class="ada-card" style="border-color:{BRS};">
                    <div style="font-size:.88rem;font-weight:700;color:{AC};">{t("endo_title")}</div>
                    <div style="color:{TXS};font-size:.77rem;margin-top:.3rem;line-height:1.6;">
                        ΔH° = {dH:.3f} kJ/mol<br>{t("endo_body")}</div></div>""",unsafe_allow_html=True)

            if dS>0:
                st.markdown(f"""<div class="ada-card">
                    <div style="font-size:.88rem;font-weight:700;color:#34d399;">{t("entropy_pos")}</div>
                    <div style="color:{TXS};font-size:.77rem;margin-top:.3rem;">ΔS° = {dS:.3f} J/mol·K<br>{t("entropy_pos_body")}</div></div>""",unsafe_allow_html=True)
            else:
                st.markdown(f"""<div class="ada-card">
                    <div style="font-size:.88rem;font-weight:700;color:{TXM};">{t("entropy_neg")}</div>
                    <div style="color:{TXS};font-size:.77rem;margin-top:.3rem;">ΔS° = {dS:.3f} J/mol·K<br>{t("entropy_neg_body")}</div></div>""",unsafe_allow_html=True)

            # auto interp
            st.markdown(f"<br><b style='color:{TXS};font-family:{ff};'>{t('auto_interp')}</b>",unsafe_allow_html=True)
            tlines=[
                f"• {'Exothermic' if dH<0 else 'Endothermic'}: ΔH° = {dH:.3f} kJ/mol",
                f"• {'Increased' if dS>0 else 'Decreased'} entropy: ΔS° = {dS:.3f} J/mol·K",
                f"• ΔG° at T₁ = {df_th['dG'].iloc[0]:.3f} kJ/mol ({'spontaneous' if df_th['dG'].iloc[0]<0 else 'non-spontaneous'})",
                f"• Van't Hoff R² = {r2_th:.4f}",
            ]
            if abs(dH)<20: tlines.append("• |ΔH°| < 20 kJ/mol → Physisorption")
            elif abs(dH)<40: tlines.append("• 20–40 kJ/mol → Mixed physisorption/chemisorption")
            else: tlines.append("• |ΔH°| > 40 kJ/mol → Chemisorption")
            st.markdown(f"""<div class="interp-box">
                {"".join([f'<div class="interp-line">{l}</div>' for l in tlines])}</div>""",unsafe_allow_html=True)

            st.markdown(f"<br><b style='color:{TXS};font-size:.84rem;font-family:{ff};'>{t('dg_table')}</b>",unsafe_allow_html=True)
            dgd=df_th[["T_C","T_K","Kc","lnKc","dG"]].copy().round(4)
            dgd.columns=["T(°C)","T(K)","Kc","ln Kc","ΔG°(kJ/mol)"]
            st.dataframe(dgd,use_container_width=True,height=195)
            if (df_th["dG"]<0).all(): st.success(t("spontaneous"))
            else: st.warning(t("non_spontaneous"))

        st.session_state["thermo_results"]={"delta_H":dH,"delta_S":dS,"r2":r2_th,"df":dgd}
    except Exception as e: st.error(str(e))

# ─────────────────────────────────────────────────────────────────────────────
#  REPORT
# ─────────────────────────────────────────────────────────────────────────────
with tab_rep:
    st.markdown(f"""
    <div class="ada-section"><div class="ada-icon">📄</div>
    <div><div class="ada-title">{t("report_header")}</div>
    <div class="ada-desc">{t("report_desc")}</div></div></div>""",unsafe_allow_html=True)

    rp1,rp2=st.columns([2,1])
    with rp1:
        exp_nm=st.text_input(t("exp_name"),value="Adsorption Study - Sample 01")
        rsr=st.text_input(t("researcher"),value="")
        nts=st.text_area(t("notes_label"),value="",height=72)
    with rp2:
        st.markdown(f"""
        <div class="ada-card" style="margin-top:.3rem;line-height:2.1;">
            <div style="font-weight:700;color:{AC};font-size:.83rem;margin-bottom:.45rem;">{t("export_includes")}</div>
            <div style="color:{TXS};font-size:.76rem;">
                ✅ Cover sheet with study metadata<br>✅ Isotherm analysis — 6 models, styled<br>
                ✅ Favorability indicators (RL, 1/n, E, ns, bT)<br>✅ Kinetics — 4 models + mechanism labels<br>
                ✅ Thermodynamics — ΔH°, ΔS°, ΔG° table<br>✅ R², RMSE, χ² color-coded<br>
                ✅ Full scientific interpretation per model<br>✅ Raw experimental data sheets
            </div>
        </div>""",unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    if st.button(t("generate_btn"),use_container_width=True):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            AR = is_ar  # bilingual flag

            # ── colour palette ──────────────────────────────────────────
            CB="1E1B4B"; CP="7C3AED"; CL="EDE9FE"; CW="FFFFFF"; CA="F5F3FF"
            CGL="DCFCE7"; CY="D97706"; CBL="2563EB"; CGR="6B7280"
            C_OK="059669"; C_BAD="DC2626"

            _thin=Side(style="thin",color="D1D5DB")
            _brd=Border(left=_thin,right=_thin,top=_thin,bottom=_thin)

            def _f(sz=10,bold=False,color="1F2937",italic=False):
                return XFont(name="Calibri",size=sz,bold=bold,color=color,italic=italic)
            def _fill(c): return PatternFill("solid",fgColor=c)
            def _ac(h="center",v="center",wrap=True,rtl=False):
                ro=2 if rtl else 1
                return Alignment(horizontal=h,vertical=v,wrapText=wrap,readingOrder=ro)
            def _al(rtl=False): return _ac("right" if rtl else "left",rtl=rtl)
            def _title(ws,text,ncols,bg=CB,sz=14,ht=38):
                ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
                c=ws["A1"]; c.value=text
                c.font=_f(sz,True,CW); c.fill=_fill(bg)
                c.alignment=_ac(rtl=AR); ws.row_dimensions[1].height=ht
            def _set_cols(ws,widths):
                for i,w in enumerate(widths):
                    ws.column_dimensions[get_column_letter(i+1)].width=w
            def _r2color(v): return C_OK if v>=0.99 else (CY if v>=0.95 else C_BAD)

            # ── bilingual label dictionary ──────────────────────────────
            LD={
                # Cover
                "cover_title":    ("AdsorpLab Pro — Scientific Adsorption Analysis Report",
                                   "أدسوربلاب برو — تقرير تحليل الأدمصاص العلمي"),
                "cover_authors":  ("by Diana Raie, Abdelbaset A. A. Diyab",
                                   "إعداد: Diana Raie، Abdelbaset A. A. Diyab"),
                "cover_sheet":    ("Cover","الغلاف"),
                "quick_summary":  ("QUICK SUMMARY","ملخص سريع"),
                "lbl_experiment": ("Experiment","اسم التجربة"),
                "lbl_researcher": ("Researcher","الباحث"),
                "lbl_notes":      ("Notes","ملاحظات"),
                "lbl_generated":  ("Generated By","أُنشئ بواسطة"),
                "lbl_date":       ("Date","التاريخ"),
                "lbl_best_iso":   ("Best Isotherm Model","أفضل نموذج أيزوثيرم"),
                "lbl_best_kin":   ("Best Kinetic Model","أفضل نموذج حركي"),
                "lbl_thermo_nat": ("Thermodynamic Nature","الطبيعة الثرموديناميكية"),
                # Isotherm sheet
                "iso_sheet":      ("Isotherm Analysis","تحليل الأيزوثيرم"),
                "iso_title":      ("ISOTHERM ANALYSIS — MODEL COMPARISON & PARAMETERS",
                                   "تحليل الأيزوثيرم — مقارنة النماذج والمعاملات"),
                "iso_params_hdr": ("MODEL PARAMETERS — DETAILED","معاملات النماذج — تفصيلية"),
                "h_model":        ("Model","النموذج"),
                "h_fav":          ("Favorability Indicator","مؤشر الملاءمة"),
                "h_interp":       ("Scientific Interpretation","التفسير العلمي"),
                "h_mech":         ("Mechanism","الميكانيكية"),
                "h_param":        ("Parameter","المعامل"),
                "h_value":        ("Value","القيمة"),
                "h_unit":         ("Unit / Notes","الوحدة / ملاحظات"),
                # Kinetics sheet
                "kin_sheet":      ("Kinetics Analysis","تحليل الحركية"),
                "kin_title":      ("KINETICS ANALYSIS — MODEL COMPARISON & PARAMETERS",
                                   "تحليل الحركية — مقارنة النماذج والمعاملات"),
                "kin_params_hdr": ("KINETIC PARAMETERS — DETAILED","المعاملات الحركية — تفصيلية"),
                # Thermodynamics sheet
                "thermo_sheet":   ("Thermodynamics","الثرموديناميكا"),
                "thermo_title":   ("THERMODYNAMIC ANALYSIS — Van\u2019t Hoff",
                                   "التحليل الثرموديناميكي — فانت هوف"),
                "h_description":  ("Description","الوصف"),
                "h_interpretation":("Interpretation","التفسير"),
                "dg_section":     ("GIBBS FREE ENERGY vs TEMPERATURE (\u0394G\u00b0)",
                                   "طاقة جيبس الحرة مقابل درجة الحرارة (\u0394G\u00b0)"),
                # Raw data sheets
                "raw_iso":        ("Isotherm Raw Data","البيانات الخام — الأيزوثيرم"),
                "raw_kin":        ("Kinetics Raw Data","البيانات الخام — الحركية"),
                # Thermo param labels
                "dH_desc":        ("Enthalpy change of adsorption","تغيّر المحتوى الحراري للأدمصاص"),
                "dS_desc":        ("Entropy change at solid–liquid interface","تغيّر الإنتروبي عند واجهة صلب-سائل"),
                "r2_desc":        ("Linearity of ln(Kc) vs 1/T","خطية ln(Kc) مقابل 1/T"),
                "type_desc":      ("|ΔH°|: <20→Physi | 20–40→Mixed | >40→Chemi",
                                   "|ΔH°|: <20→فيزيائي | 20-40→مختلط | >40→كيميائي"),
                "ads_type_lbl":   ("Adsorption Type","نوع الأدمصاص"),
                # Thermo interpretations
                "exo_interp":     ("Exothermic — heat released; spontaneous tendency increases with lower T",
                                   "طارد للحرارة — تنبعث حرارة؛ العفوية تزيد بانخفاض الحرارة"),
                "endo_interp":    ("Endothermic — heat absorbed; driven by entropy increase",
                                   "ماص للحرارة — تمتص حرارة؛ مدفوع بزيادة الإنتروبي"),
                "dS_pos":         ("Increased randomness at solid–liquid interface (+ΔS°)",
                                   "اضطراب متزايد عند الواجهة (+ΔS°)"),
                "dS_neg":         ("Decreased randomness; ordered adsorption layer",
                                   "اضطراب متناقص؛ طبقة أدمصاص منتظمة"),
                "r2_excel":       ("Excellent — thermodynamic data reliable",
                                   "ممتاز — البيانات الثرموديناميكية موثوقة"),
                "r2_good":        ("Good fit","مطابقة جيدة"),
                "r2_mod":         ("Moderate — check data consistency","متوسط — تحقق من اتساق البيانات"),
                "physi":          ("Physisorption","أدمصاص فيزيائي"),
                "mixed":          ("Mixed physi/chemisorption","أدمصاص مختلط فيزيائي/كيميائي"),
                "chemi":          ("Chemisorption","أدمصاص كيميائي"),
                "ion_exch":       ("Ion Exchange","تبادل أيوني"),
            }
            def L(k): return LD[k][1 if AR else 0]

            # ── unit notes (bilingual) ──────────────────────────────────
            _punits={
                "qₘₐₓ (mg/g)": ("mg/g","mg/g"),
                "KL (L/mg)":   ("L/mg","L/mg"),
                "RL":          ("dimensionless | 0<RL<1: favorable","بُعدي | 0<RL<1: ملائم"),
                "KF":          ("(mg/g)(L/mg)^(1/n)","(mg/g)(L/mg)^(1/n)"),
                "1/n":         ("dimensionless | <1: favorable","بُعدي | <1: ملائم"),
                "n":           ("dimensionless","بُعدي"),
                "AT (L/g)":    ("L/g","L/g"),
                "B (mg/g)":    ("J/mol","J/mol"),
                "bT (J/mol)":  ("J/mol | <40 000: physisorption","J/mol | <40 000: فيزيائي"),
                "qₘ (mg/g)":   ("mg/g","mg/g"),
                "\u03b2 (mol\u00b2/kJ\u00b2)":("mol²/kJ²","mol²/kJ²"),
                "E (kJ/mol)":  ("kJ/mol | <8:physi | 8-16:ion | >16:chemi",
                                "kJ/mol | <8:فيزيائي | 8-16:أيوني | >16:كيميائي"),
                "Type":        ("Adsorption mechanism","آلية الأدمصاص"),
                "Ks":          ("(L/mg)^ns","(L/mg)^ns"),
                "ns":          ("heterogeneity | ns=1→Langmuir","عدم تجانس | ns=1→لانجمير"),
                "Cs (mg/L)":   ("mg/L","mg/L"),
                "C_BET":       ("dimensionless BET constant","ثابت BET بُعدي"),
            }
            def PU(k): return _punits.get(k,("",""))[1 if AR else 0]

            # ── kinetics mechanism labels (bilingual) ───────────────────
            _mechs={
                "PFO":(
                    ("Pseudo-First-Order — Physical Adsorption",
                     "Rate \u221d (qe\u2212qt). R\u00b2\u22650.99: film diffusion likely rate-limiting. Indicates physisorption or reversible binding."),
                    ("من الدرجة الأولى الزائف — أدمصاص فيزيائي",
                     "المعدل \u221d (qe\u2212qt). R\u00b2\u22650.99: انتشار الفيلم محدود السرعة. يدل على الأدمصاص الفيزيائي أو الارتباط العكسي.")
                ),
                "PSO":(
                    ("Pseudo-Second-Order — Chemisorption",
                     "Rate \u221d (qe\u2212qt)\u00b2. Best fit: chemical bond formation. PSO>PFO in R\u00b2 \u2192 chemisorption dominant."),
                    ("من الدرجة الثانية الزائف — أدمصاص كيميائي",
                     "المعدل \u221d (qe\u2212qt)\u00b2. أفضل مطابقة: تكوين رابطة كيميائية. PSO>PFO في R\u00b2 \u2192 الأدمصاص الكيميائي سائد.")
                ),
                "Elovich":(
                    ("Elovich — Heterogeneous Chemisorption",
                     "\u03b1: initial rate (mg/g\u00b7min); \u03b2: desorption constant (g/mg). Applies to activated heterogeneous surfaces."),
                    ("إيلوفيتش — أدمصاص كيميائي غير متجانس",
                     "\u03b1: معدل أولي (mg/g\u00b7min)؛ \u03b2: ثابت الانفصال (g/mg). ينطبق على الأسطح غير المتجانسة.")
                ),
                "Weber-Morris":(
                    ("Weber-Morris — Intraparticle Diffusion",
                     "kid: diffusion rate constant. C\u22480 \u2192 IPD rate-limiting. C>0 \u2192 combined film+IPD mechanism."),
                    ("ويبر-موريس — انتشار داخل الجسيمات",
                     "kid: ثابت معدل الانتشار. C\u22480 \u2192 الانتشار الداخلي محدود السرعة. C>0 \u2192 آلية مركبة.")
                ),
            }
            def MK(nm): m=_mechs.get(nm,(("—","—"),("—","—"))); return m[1 if AR else 0]

            # ── favorability function (bilingual) ───────────────────────
            def _iso_fav(nm,params):
                if nm=="Langmuir":
                    RL=params.get("RL",0.5)
                    if RL<=0:
                        return (f"RL={RL:.4f} — "+("Irreversible" if not AR else "لا رجعي"),"7C3AED",
                                ("Irreversible adsorption; adsorbate strongly bound to surface." if not AR
                                 else "أدمصاص لا رجعي؛ الأدمصوص مرتبط بقوة بالسطح."))
                    elif RL<1:
                        return (f"RL={RL:.4f} — "+("\u2714 Favorable" if not AR else "\u2714 ملائم"),C_OK,
                                (f"Favorable monolayer adsorption (0<RL<1). High affinity. RL={RL:.4f}." if not AR
                                 else f"أدمصاص أحادي الطبقة ملائم (0<RL<1). تقارب عالٍ. RL={RL:.4f}."))
                    elif RL==1:
                        return (f"RL={RL:.4f} — "+("Linear" if not AR else "خطي"),CY,
                                ("Linear isotherm; proportional to concentration." if not AR
                                 else "منحنى أيزوثيرم خطي؛ الأدمصاص متناسب مع التركيز."))
                    else:
                        return (f"RL={RL:.4f} — "+("\u2718 Unfavorable" if not AR else "\u2718 غير ملائم"),C_BAD,
                                (f"Unfavorable (RL>1). Low affinity; consider adjusting pH/ionic strength." if not AR
                                 else f"غير ملائم (RL>1). تقارب منخفض؛ يُنصح بتعديل الرقم الهيدروجيني/قوة أيونية."))
                elif nm=="Freundlich":
                    inv_n=params.get("1/n",1.0)
                    if inv_n<1:
                        return (f"1/n={inv_n:.4f} — "+("\u2714 Favorable" if not AR else "\u2714 ملائم"),C_OK,
                                (f"Favorable heterogeneous adsorption (1/n<1). Multi-site energy distribution." if not AR
                                 else f"أدمصاص غير متجانس ملائم (1/n<1). توزيع طاقي لمواقع الارتباط."))
                    elif inv_n>1:
                        return (f"1/n={inv_n:.4f} — "+("\u2718 Unfavorable" if not AR else "\u2718 غير ملائم"),C_BAD,
                                (f"Unfavorable (1/n>1). Anti-Langmuir cooperative behaviour." if not AR
                                 else f"غير ملائم (1/n>1). سلوك تعاوني معاكس للانجمير."))
                    else:
                        return (f"1/n={inv_n:.4f} — "+("Linear" if not AR else "خطي"),CY,
                                ("Linear Freundlich; homogeneous surface energy." if not AR
                                 else "فرندليش خطي؛ طاقة سطح متجانسة."))
                elif nm=="D-R":
                    E=params.get("E (kJ/mol)",0); at=params.get("Type","")
                    if "Physi" in at:
                        return (f"E={E:.2f} kJ/mol — "+("Physisorption" if not AR else "أدمصاص فيزيائي"),CBL,
                                (f"E<8 kJ/mol: van der Waals forces dominant. Reversible low-energy adsorption." if not AR
                                 else f"E<8 kJ/mol: قوى فان ديرفالس سائدة. أدمصاص فيزيائي عكسي منخفض الطاقة."))
                    elif "Ion" in at:
                        return (f"E={E:.2f} kJ/mol — "+("Ion Exchange" if not AR else "تبادل أيوني"),CY,
                                (f"8\u2264E<16 kJ/mol: Ion exchange mechanism; intermediate binding energy." if not AR
                                 else f"8\u2264E<16 kJ/mol: آلية التبادل الأيوني؛ طاقة ارتباط متوسطة."))
                    else:
                        return (f"E={E:.2f} kJ/mol — "+("Chemisorption" if not AR else "أدمصاص كيميائي"),C_BAD,
                                (f"E\u226516 kJ/mol: Chemical bond formation. Strong, potentially irreversible." if not AR
                                 else f"E\u226516 kJ/mol: تكوين رابطة كيميائية. أدمصاص قوي قد يكون لا رجعياً."))
                elif nm=="Sips":
                    ns=params.get("ns",1.0)
                    if abs(ns-1)<0.15:
                        return (f"ns={ns:.4f} — "+("\u2248Langmuir" if not AR else "\u2248لانجمير"),CY,
                                (f"ns\u22481: Nearly homogeneous surface (approaches Langmuir). ns={ns:.4f}." if not AR
                                 else f"ns\u22481: سطح متجانس تقريباً (يقترب من لانجمير). ns={ns:.4f}."))
                    else:
                        return (f"ns={ns:.4f} — "+("Heterogeneous" if not AR else "غير متجانس"),C_OK,
                                (f"ns\u22601: Heterogeneous surface; distributed binding energies. ns={ns:.4f}." if not AR
                                 else f"ns\u22601: سطح غير متجانس؛ توزيع طاقي لمواقع الارتباط. ns={ns:.4f}."))
                elif nm=="Temkin":
                    bT=params.get("bT (J/mol)",0)
                    lbl=("Physisorption" if not AR else "أدمصاص فيزيائي") if bT<40000 else ("Chemisorption" if not AR else "أدمصاص كيميائي")
                    clr=CBL if bT<40000 else C_BAD
                    interp=(f"bT={bT:.0f} J/mol. {'<40 kJ/mol: physical adsorption dominant.' if bT<40000 else '≥40 kJ/mol: chemical interactions dominant.'}"
                            if not AR else
                            f"bT={bT:.0f} J/mol. {'<40 kJ/mol: الأدمصاص الفيزيائي سائد.' if bT<40000 else '≥40 kJ/mol: التفاعلات الكيميائية سائدة.'}")
                    return f"bT={bT:.0f} J/mol — {lbl}",clr,interp
                elif nm=="BET":
                    C_b=params.get("C_BET",0)
                    return (f"C_BET={C_b:.2f}",CP,
                            (f"BET constant C={C_b:.2f}. Multilayer model. C>>1: strong adsorbate–surface interaction." if not AR
                             else f"ثابت BET C={C_b:.2f}. نموذج متعدد الطبقات. C>>1: تفاعل قوي أدمصوص-سطح."))
                return "—","1F2937","—"

            wb=Workbook(); wb.remove(wb.active)

            # ══════════════════════════════════════════════════════════════
            # SHEET 1 — COVER
            # ══════════════════════════════════════════════════════════════
            wc=wb.create_sheet(L("cover_sheet")); _set_cols(wc,[32,46,22])
            _title(wc,L("cover_title"),3,CB,16,42)
            wc.merge_cells("A2:C2"); s=wc["A2"]
            s.value=L("cover_authors")
            s.font=_f(11,False,CW,italic=True); s.fill=_fill(CP); s.alignment=_ac(rtl=AR)
            wc.row_dimensions[2].height=22

            meta=[(L("lbl_experiment"),exp_nm),(L("lbl_researcher"),rsr or "—"),
                  (L("lbl_notes"),nts or "—"),(L("lbl_generated"),"AdsorpLab Pro v2.0"),
                  (L("lbl_date"),pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"))]
            for i,(k,v) in enumerate(meta):
                r=i+4; wc.row_dimensions[r].height=21
                a=wc.cell(r,1,k); a.font=_f(10,True,CP); a.fill=_fill(CL); a.alignment=_al(AR); a.border=_brd
                b=wc.cell(r,2,v); b.font=_f(10); b.fill=_fill(CW); b.alignment=_al(AR); b.border=_brd
                wc.cell(r,3,"").fill=_fill(CW); wc.cell(r,3).border=_brd

            r=10; wc.row_dimensions[r].height=26; wc.merge_cells(f"A{r}:C{r}")
            s=wc[f"A{r}"]; s.value=L("quick_summary"); s.font=_f(12,True,CW); s.fill=_fill(CP); s.alignment=_ac(rtl=AR)

            smry=[]
            if "iso_results" in st.session_state:
                bi=st.session_state.get("iso_best","—"); bv=st.session_state["iso_results"].get(bi,{}).get("r2",0)
                smry.append((L("lbl_best_iso"),bi,f"R\u00b2 = {bv:.4f}",C_OK if bv>=0.99 else CY))
            if "kin_results" in st.session_state:
                bk2=st.session_state.get("kin_best","—"); bkv=st.session_state["kin_results"].get(bk2,{}).get("r2",0)
                smry.append((L("lbl_best_kin"),bk2,f"R\u00b2 = {bkv:.4f}",C_OK if bkv>=0.99 else CY))
            if "thermo_results" in st.session_state:
                tr2=st.session_state["thermo_results"]; dHs=tr2["delta_H"]
                nat=(("Exothermic" if not AR else "طارد للحرارة") if dHs<0 else ("Endothermic" if not AR else "ماص للحرارة"))
                aty=L("physi") if abs(dHs)<20 else (L("mixed") if abs(dHs)<40 else L("chemi"))
                smry.append((L("lbl_thermo_nat"),f"{nat} — {aty}",f"\u0394H\u00b0 = {dHs:.3f} kJ/mol",C_OK if dHs<0 else C_BAD))
            for i,(label,val,stat,vc) in enumerate(smry):
                r=11+i; wc.row_dimensions[r].height=21
                a=wc.cell(r,1,label); a.font=_f(10,True); a.fill=_fill(CA); a.alignment=_al(AR); a.border=_brd
                b=wc.cell(r,2,val); b.font=_f(10,True,vc); b.fill=_fill(CW); b.alignment=_al(AR); b.border=_brd
                c2=wc.cell(r,3,stat); c2.font=_f(9,color=CGR); c2.fill=_fill(CW); c2.alignment=_ac(rtl=AR); c2.border=_brd

            # ══════════════════════════════════════════════════════════════
            # SHEET 2 — ISOTHERM ANALYSIS
            # ══════════════════════════════════════════════════════════════
            if "iso_results" in st.session_state:
                ir=st.session_state["iso_results"]; bi=st.session_state.get("iso_best","—")
                wi=wb.create_sheet(L("iso_sheet")); _set_cols(wi,[20,10,10,10,32,44])
                _title(wi,L("iso_title"),6,CB,14,36)
                hdrs=[L("h_model"),"R\u00b2","RMSE","\u03c7\u00b2",L("h_fav"),L("h_interp")]
                for ci,h in enumerate(hdrs):
                    ce=wi.cell(2,ci+1,h); ce.font=_f(10,True,CW); ce.fill=_fill(CP); ce.alignment=_ac(rtl=AR); ce.border=_brd
                wi.row_dimensions[2].height=28

                srt=sorted(ir.items(),key=lambda x:-x[1]["r2"])
                for ri,(nm,res) in enumerate(srt):
                    row=ri+3; alt=ri%2==1; ib=(nm==bi)
                    bg=CGL if ib else (CA if alt else CW)
                    wi.row_dimensions[row].height=20
                    fl,fc,fi=_iso_fav(nm,res["params"])
                    r2v=res["r2"]
                    data_row=[
                        (("★ " if ib else "")+nm, CP if ib else "1F2937", ib, _al(AR)),
                        (round(r2v,4),   _r2color(r2v), True,  _ac(rtl=AR)),
                        (round(res["rmse"],4),"374151",False,_ac(rtl=AR)),
                        (round(res["chi2"],4),"374151",False,_ac(rtl=AR)),
                        (fl, fc, True, _al(AR)),
                        (fi,"374151",False,_al(AR)),
                    ]
                    for ci,(val,clr,bold,algn) in enumerate(data_row):
                        ce=wi.cell(row,ci+1,val); ce.font=_f(10,bold,clr)
                        ce.fill=_fill(bg); ce.border=_brd; ce.alignment=algn

                # Parameters sub-table
                pr=len(srt)+4; wi.row_dimensions[pr].height=26
                wi.merge_cells(f"A{pr}:F{pr}"); s=wi[f"A{pr}"]
                s.value=L("iso_params_hdr"); s.font=_f(11,True,CW); s.fill=_fill(CP); s.alignment=_ac(rtl=AR)
                pr+=1
                for ci,h in enumerate([L("h_model"),L("h_param"),L("h_value"),L("h_unit")]):
                    ce=wi.cell(pr,ci+1,h); ce.font=_f(10,True,CW); ce.fill=_fill(CP); ce.alignment=_ac(rtl=AR); ce.border=_brd
                wi.row_dimensions[pr].height=22; pr+=1
                for ri2,(nm,res) in enumerate(srt):
                    bg=CA if ri2%2 else CW
                    for pk,pv in res["params"].items():
                        wi.row_dimensions[pr].height=18
                        for ci2,(val,bold,clr,algn) in enumerate([
                            (nm,True,CP,_al(AR)),(pk,True,"1F2937",_al(AR)),
                            (f"{pv:.6f}" if isinstance(pv,float) else str(pv),False,"1D4ED8",_ac(rtl=AR)),
                            (PU(pk),False,CGR,_al(AR))]):
                            ce=wi.cell(pr,ci2+1,val); ce.font=_f(9,bold,clr)
                            ce.fill=_fill(bg); ce.border=_brd; ce.alignment=algn
                        pr+=1

            # ══════════════════════════════════════════════════════════════
            # SHEET 3 — KINETICS ANALYSIS
            # ══════════════════════════════════════════════════════════════
            if "kin_results" in st.session_state:
                kr=st.session_state["kin_results"]; bk=st.session_state.get("kin_best","—")
                wk=wb.create_sheet(L("kin_sheet")); _set_cols(wk,[22,10,10,10,32,44])
                _title(wk,L("kin_title"),6,CB,14,36)
                for ci,h in enumerate([L("h_model"),"R\u00b2","RMSE","\u03c7\u00b2",L("h_mech"),L("h_interp")]):
                    ce=wk.cell(2,ci+1,h); ce.font=_f(10,True,CW); ce.fill=_fill(CP); ce.alignment=_ac(rtl=AR); ce.border=_brd
                wk.row_dimensions[2].height=28

                srtk=sorted(kr.items(),key=lambda x:-x[1]["r2"])
                for ri,(nm,res) in enumerate(srtk):
                    row=ri+3; alt=ri%2==1; ib=(nm==bk); bg=CGL if ib else (CA if alt else CW)
                    wk.row_dimensions[row].height=20
                    mlab,minterp=MK(nm)
                    r2v=res["r2"]
                    for ci,(val,clr,bold,algn) in enumerate([
                        (("★ " if ib else "")+nm, CP if ib else "1F2937", ib, _al(AR)),
                        (round(r2v,4), _r2color(r2v), True, _ac(rtl=AR)),
                        (round(res["rmse"],4),"374151",False,_ac(rtl=AR)),
                        (round(res["chi2"],4),"374151",False,_ac(rtl=AR)),
                        (mlab, CP, True, _al(AR)),
                        (minterp,"374151",False,_al(AR)),
                    ]):
                        ce=wk.cell(row,ci+1,val); ce.font=_f(10,bold,clr)
                        ce.fill=_fill(bg); ce.border=_brd; ce.alignment=algn

                pr=len(srtk)+4; wk.row_dimensions[pr].height=26
                wk.merge_cells(f"A{pr}:F{pr}"); s=wk[f"A{pr}"]
                s.value=L("kin_params_hdr"); s.font=_f(11,True,CW); s.fill=_fill(CP); s.alignment=_ac(rtl=AR)
                pr+=1
                for ci,h in enumerate([L("h_model"),L("h_param"),L("h_value"),L("h_unit")]):
                    ce=wk.cell(pr,ci+1,h); ce.font=_f(10,True,CW); ce.fill=_fill(CP); ce.alignment=_ac(rtl=AR); ce.border=_brd
                wk.row_dimensions[pr].height=22; pr+=1
                _ku={
                    "k1 (1/min)":       ("1/min","1/min"),
                    "qe (mg/g)":        ("mg/g","mg/g"),
                    "k2 (g/mg\u00b7min)":("g/mg\u00b7min","g/mg\u00b7min"),
                    "\u03b1 (mg/g\u00b7min)":("mg/g\u00b7min — initial rate","mg/g\u00b7min — معدل أولي"),
                    "\u03b2 (g/mg)":    ("g/mg — desorption const.","g/mg — ثابت الانفصال"),
                    "kid (mg/g\u00b7min^0.5)":("mg/g\u00b7min^0.5","mg/g\u00b7min^0.5"),
                    "C (mg/g)":         ("mg/g — boundary layer","mg/g — سُمك طبقة الحد"),
                }
                for ri2,(nm,res) in enumerate(srtk):
                    bg=CA if ri2%2 else CW
                    for pk,pv in res["params"].items():
                        wk.row_dimensions[pr].height=18
                        ku_val=_ku.get(pk,("",""))[1 if AR else 0]
                        for ci2,(val,bold,clr,algn) in enumerate([
                            (nm,True,CP,_al(AR)),(pk,True,"1F2937",_al(AR)),
                            (f"{pv:.6f}",False,"1D4ED8",_ac(rtl=AR)),(ku_val,False,CGR,_al(AR))]):
                            ce=wk.cell(pr,ci2+1,val); ce.font=_f(9,bold,clr)
                            ce.fill=_fill(bg); ce.border=_brd; ce.alignment=algn
                        pr+=1

            # ══════════════════════════════════════════════════════════════
            # SHEET 4 — THERMODYNAMICS
            # ══════════════════════════════════════════════════════════════
            if "thermo_results" in st.session_state:
                tr=st.session_state["thermo_results"]
                wt=wb.create_sheet(L("thermo_sheet")); _set_cols(wt,[24,16,34,30,10])
                _title(wt,L("thermo_title"),5,CB,14,36)
                for ci,h in enumerate([L("h_param"),L("h_value"),L("h_description"),L("h_interpretation"),""]):
                    ce=wt.cell(2,ci+1,h); ce.font=_f(10,True,CW); ce.fill=_fill(CP); ce.alignment=_ac(rtl=AR); ce.border=_brd
                wt.row_dimensions[2].height=26

                dHv=tr["delta_H"]; dSv=tr["delta_S"]; r2v=tr["r2"]
                aty=L("physi") if abs(dHv)<20 else (L("mixed") if abs(dHv)<40 else L("chemi"))
                tp=[
                    ("\u0394H\u00b0 (kJ/mol)",f"{dHv:.4f}",L("dH_desc"),
                     L("exo_interp") if dHv<0 else L("endo_interp"), C_OK if dHv<0 else C_BAD),
                    ("\u0394S\u00b0 (J/mol\u00b7K)",f"{dSv:.4f}",L("dS_desc"),
                     L("dS_pos") if dSv>0 else L("dS_neg"), C_OK if dSv>0 else CGR),
                    ("Van\u2019t Hoff R\u00b2",f"{r2v:.4f}",L("r2_desc"),
                     L("r2_excel") if r2v>=0.99 else (L("r2_good") if r2v>=0.95 else L("r2_mod")),
                     C_OK if r2v>=0.99 else (CY if r2v>=0.95 else C_BAD)),
                    (L("ads_type_lbl"),aty,L("type_desc"),f"|\u0394H\u00b0| = {abs(dHv):.1f} kJ/mol",CP),
                ]
                for ri,(p,v,d,interp,vc) in enumerate(tp):
                    row=ri+3; bg=CA if ri%2 else CW; wt.row_dimensions[row].height=22
                    for ci2,(val,bold,clr,algn) in enumerate([
                        (p,True,"1F2937",_al(AR)),(v,True,"1D4ED8",_ac(rtl=AR)),
                        (d,False,CGR,_al(AR)),(interp,True,vc,_al(AR)),("",False,CGR,_ac())]):
                        ce=wt.cell(row,ci2+1,val); ce.font=_f(10,bold,clr)
                        ce.fill=_fill(bg); ce.border=_brd; ce.alignment=algn

                pr=8; wt.row_dimensions[pr].height=26; wt.merge_cells(f"A{pr}:E{pr}")
                s=wt[f"A{pr}"]; s.value=L("dg_section"); s.font=_f(11,True,CW)
                s.fill=_fill(CP); s.alignment=_ac(rtl=AR); pr+=1
                dg=tr["df"]
                for ci,col in enumerate(dg.columns):
                    ce=wt.cell(pr,ci+1,col); ce.font=_f(10,True,CW); ce.fill=_fill(CP); ce.alignment=_ac(rtl=AR); ce.border=_brd
                wt.row_dimensions[pr].height=22; pr+=1
                for ri2,row_d in dg.iterrows():
                    bg=CA if ri2%2 else CW; wt.row_dimensions[pr].height=18
                    for ci2,val in enumerate(row_d):
                        try: vf=round(float(val),4)
                        except: vf=val
                        ce=wt.cell(pr,ci2+1,vf)
                        is_dg="dG" in str(dg.columns[ci2]).lower() or "\u0394G" in str(dg.columns[ci2])
                        ce.font=_f(9,is_dg,C_OK if (is_dg and isinstance(vf,(int,float)) and vf<0) else
                                           (C_BAD if (is_dg and isinstance(vf,(int,float)) and vf>=0) else "374151"))
                        ce.fill=_fill(bg); ce.border=_brd; ce.alignment=_ac(rtl=AR)
                    pr+=1

            # ══════════════════════════════════════════════════════════════
            # SHEET 5 & 6 — RAW DATA
            # ══════════════════════════════════════════════════════════════
            for sheet_nm,ss_key in [(L("raw_iso"),"iso_df"),(L("raw_kin"),"kin_df")]:
                if ss_key in st.session_state:
                    df_src=st.session_state[ss_key]; wd=wb.create_sheet(sheet_nm)
                    _set_cols(wd,[18]*len(df_src.columns))
                    wd.row_dimensions[1].height=26
                    for ci,col in enumerate(df_src.columns):
                        ce=wd.cell(1,ci+1,col); ce.font=_f(10,True,CW)
                        ce.fill=_fill(CP); ce.alignment=_ac(rtl=AR); ce.border=_brd
                    for ri2,row_d in df_src.iterrows():
                        bg=CA if ri2%2 else CW; wd.row_dimensions[ri2+2].height=18
                        for ci2,val in enumerate(row_d):
                            try: vf=round(float(val),4)
                            except: vf=val
                            ce=wd.cell(ri2+2,ci2+1,vf); ce.font=_f(9)
                            ce.fill=_fill(bg); ce.border=_brd; ce.alignment=_ac(rtl=AR)

            buf=io.BytesIO(); wb.save(buf); buf.seek(0)
            safe=exp_nm.replace(" ","_").replace("/","-")[:28]
            st.download_button(t("download_btn"),buf.getvalue(),f"AdsorpLab_{safe}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
            st.success(t("report_success"))
        except Exception as e: st.error(str(e))

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown(f"""
    <div style="font-size:.93rem;font-weight:700;color:{TX};margin-bottom:.75rem;
        padding-bottom:.45rem;border-bottom:1px solid {BR};font-family:{ff};direction:{fd};">
        {t("session_results")}
    </div>""",unsafe_allow_html=True)

    has=any(k in st.session_state for k in["iso_results","kin_results","thermo_results"])
    if not has:
        st.markdown(f"""<div style="text-align:center;padding:2.2rem;color:{TXM};font-family:{ff};">
            <div style="font-size:2rem;margin-bottom:.5rem;">📭</div>
            <div style="font-size:.87rem;">{t("no_results")}</div></div>""",unsafe_allow_html=True)
    else:
        rd1,rd2,rd3=st.columns(3)
        for col,key,lbl in [(rd1,"iso_results",t("iso_models_lbl")),(rd2,"kin_results",t("kin_models_lbl"))]:
            with col:
                if key in st.session_state:
                    st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{lbl}</b>",unsafe_allow_html=True)
                    bk=st.session_state.get(key.replace("results","best"),"")
                    for nm,res in st.session_state[key].items():
                        ib=(nm==bk)
                        st.markdown(f"""
                        <div style="display:flex;justify-content:space-between;align-items:center;
                            padding:.32rem .65rem;border-radius:8px;margin-bottom:.18rem;
                            background:{'rgba(34,197,94,0.08)' if ib else CARD};
                            border:1px solid {SUCBR if ib else BR};">
                            <span style="color:{TX};font-size:.8rem;font-weight:{'700' if ib else '400'};font-family:{ff};">
                                {"★ " if ib else ""}{nm}</span>
                            <span style="color:{'#34d399' if ib else AC};font-family:'JetBrains Mono',monospace;font-size:.8rem;">
                                {res['r2']:.4f}</span>
                        </div>""",unsafe_allow_html=True)
        with rd3:
            if "thermo_results" in st.session_state:
                tr=st.session_state["thermo_results"]
                st.markdown(f"<b style='color:{TXS};font-family:{ff};'>{t('thermo_lbl')}</b>",unsafe_allow_html=True)
                for lbl,val in [("ΔH°",f"{tr['delta_H']:.3f} kJ/mol"),("ΔS°",f"{tr['delta_S']:.3f} J/mol·K"),("R²",f"{tr['r2']:.4f}")]:
                    st.markdown(f"""
                    <div style="display:flex;justify-content:space-between;padding:.32rem .65rem;
                        border-radius:8px;margin-bottom:.18rem;background:{CARD};border:1px solid {BR};">
                        <span style="color:{TXM};font-size:.8rem;font-family:{ff};">{lbl}</span>
                        <span style="color:{AC};font-family:'JetBrains Mono',monospace;font-size:.8rem;">{val}</span>
                    </div>""",unsafe_allow_html=True)

# ─── Footer ──────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="ada-footer">
    🔬 AdsorpLab Pro v2.0 &nbsp;·&nbsp;
    Langmuir · Freundlich · Temkin · D-R · Sips · BET &nbsp;·&nbsp;
    PFO · PSO · Elovich · Weber-Morris · Arrhenius &nbsp;·&nbsp;
    R² · RMSE · χ² · Linear Transforms · Auto-Interpretation<br>
    <span style="font-size:.72rem;color:{AC};font-weight:600;letter-spacing:.3px;">
        Developed by Diana Raie &amp; Abdelbaset A. A. Diyab
    </span>
</div>""",unsafe_allow_html=True)
